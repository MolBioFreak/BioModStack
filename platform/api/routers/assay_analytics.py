from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import statistics
from typing import Any, Dict, Iterable, List, Optional, Tuple
import zipfile
import xml.etree.ElementTree as ET

import numpy as np
from fastapi import APIRouter, File, Form, HTTPException, UploadFile, Query
from fastapi.responses import Response
from pydantic import BaseModel, ConfigDict, Field
from scipy import stats
from scipy.signal import find_peaks, peak_widths, savgol_filter
from sqlalchemy.exc import SQLAlchemyError

from services.assay_tool_integrations import assay_tool_registry, tools_by_category
from services.assay_analytical_store import (
    AnalyticalStoreUnavailable,
    DB_SERVICE_OFFLINE_MESSAGE,
    analytical_store_status,
    list_analytical_datasets,
    load_analytical_dataset,
    list_qpcr_imports,
    load_qpcr_import,
    persist_qpcr_import_response,
)
from services.assay_chrom_persistence import persist_empower_import

router = APIRouter()


_DB_DEGRADED_EXCEPTIONS = (AnalyticalStoreUnavailable, SQLAlchemyError, OSError, ConnectionError)


def _db_service_degraded_detail(exc: BaseException) -> Dict[str, Any]:
    return {
        "component": "db-service",
        "service_id": "bms-db-service",
        "degraded_by": "bms-db-service",
        "display_name": "BMS DB service",
        "offline_message": DB_SERVICE_OFFLINE_MESSAGE,
        "message": str(exc),
        "operator_action": "Start BMS DB service from the top bar or run `bms db-service start`.",
    }


def _raise_db_service_degraded(exc: BaseException) -> None:
    raise HTTPException(status_code=503, detail=_db_service_degraded_detail(exc)) from exc

# BioModStack-native assay analytics router.
# It intentionally mirrors the prototype analysis-module API paths under
# /api/assay-analytics so the migrated React panels do not point at a defunct
# standalone parser service.


def _finite_float(value: Any, default: Optional[float] = None) -> Optional[float]:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return default
    if not math.isfinite(result):
        return default
    return result


def _json_clean(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): _json_clean(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_clean(v) for v in value]
    if isinstance(value, tuple):
        return [_json_clean(v) for v in value]
    if isinstance(value, np.ndarray):
        return _json_clean(value.tolist())
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, (np.floating, float)):
        value = float(value)
        return value if math.isfinite(value) else None
    if isinstance(value, (np.integer, int)):
        return int(value)
    return value


def _mean(values: Iterable[float]) -> float:
    vals = [float(v) for v in values]
    return float(np.mean(vals)) if vals else 0.0


def _std(values: Iterable[float], ddof: int = 1) -> float:
    vals = [float(v) for v in values]
    if len(vals) <= ddof:
        return 0.0
    return float(np.std(vals, ddof=ddof))


def _cv_percent(values: Iterable[float]) -> float:
    vals = [float(v) for v in values if _finite_float(v) is not None]
    if not vals:
        return 0.0
    m = _mean(vals)
    return abs(_std(vals) / m * 100.0) if m else 0.0


def _first_text(mapping: Dict[str, Any], aliases: Iterable[str]) -> Optional[str]:
    for alias in aliases:
        for key in _alias_keys(str(alias)):
            value = mapping.get(key)
            if value is None:
                continue
            text = str(value).strip()
            if text:
                return text
    return None


def _require_text(mapping: Dict[str, Any], aliases: Iterable[str], field_name: str, row_idx: int, context: str) -> str:
    text = _first_text(mapping, aliases)
    if text is None:
        raise HTTPException(status_code=400, detail=f"{context} row {row_idx} missing required {field_name} metadata")
    return text


def _require_finite(mapping: Dict[str, Any], aliases: Iterable[str], field_name: str, row_idx: int, context: str) -> float:
    for alias in aliases:
        value = mapping.get(alias)
        finite = _finite_float(value)
        if finite is not None:
            return finite
    raise HTTPException(status_code=400, detail=f"{context} row {row_idx} missing required finite {field_name} value")


def _required_ids(values: List[str], count: int, field_name: str) -> List[str]:
    ids = [str(value).strip() for value in values]
    if len(ids) != count or any(not value for value in ids):
        raise HTTPException(status_code=400, detail=f"{field_name} must provide one non-empty real identifier for each sample value")
    return ids


def _linear_fit(x: List[float], y: List[float]) -> Dict[str, float]:
    if len(x) != len(y) or len(x) < 2:
        raise HTTPException(status_code=400, detail="At least two paired finite points are required")
    slope, intercept, r_value, p_value, stderr = stats.linregress(x, y)
    fitted = [slope * xi + intercept for xi in x]
    residuals = [yi - fi for yi, fi in zip(y, fitted)]
    return _json_clean(
        {
            "slope": slope,
            "intercept": intercept,
            "r_squared": r_value**2,
            "p_value": p_value,
            "std_err": stderr,
            "residual_std": _std(residuals),
        }
    )


class StandardCurvePoint(BaseModel):
    cq: float
    quantity: float


class StandardCurveRequest(BaseModel):
    points: List[StandardCurvePoint] = Field(default_factory=list)
    concentrations: List[float] = Field(default_factory=list)
    cq_values: List[float] = Field(default_factory=list)
    gene: Optional[str] = None
    unit: Optional[str] = None
    log_base: float = 10.0


class QuantifyRequest(BaseModel):
    slope: Optional[float] = None
    intercept: Optional[float] = None
    cq_values: List[float] = Field(default_factory=list)
    std_concentrations: List[float] = Field(default_factory=list)
    std_cq_values: List[float] = Field(default_factory=list)
    sample_cq_values: List[float] = Field(default_factory=list)
    sample_ids: List[str] = Field(default_factory=list)
    unit: str = "quantity"
    log_base: float = 10.0


class CqRow(BaseModel):
    sample: str
    gene: str
    cq: float
    group: str


class DeltaCqRequest(BaseModel):
    rows: List[CqRow] = Field(alias="data", default_factory=list)
    reference_genes: List[str]
    target_genes: List[str]

    model_config = ConfigDict(populate_by_name=True)


class DeltaDeltaCqRequest(DeltaCqRequest):
    control_group: str


class AnovaDunnettRequest(BaseModel):
    # Array payloads must include one explicit non-empty group_names entry for each
    # group plus an explicit control_group. BMS must not invent "Control"/"Group_N"
    # labels for statistical outputs.
    groups: Any
    group_names: List[str] = Field(default_factory=list)
    control_group: str
    alpha: float = 0.05


def _standard_curve_points(request: StandardCurveRequest) -> List[StandardCurvePoint]:
    if request.points:
        return [p for p in request.points if p.quantity > 0 and math.isfinite(p.quantity) and math.isfinite(p.cq)]
    if len(request.concentrations) != len(request.cq_values):
        raise HTTPException(status_code=400, detail="concentrations and cq_values must have the same length")
    points: List[StandardCurvePoint] = []
    for quantity, cq in zip(request.concentrations, request.cq_values):
        q = _finite_float(quantity)
        c = _finite_float(cq)
        if q is not None and q > 0 and c is not None:
            points.append(StandardCurvePoint(quantity=q, cq=c))
    return points


@router.post("/analysis/qpcr/standard-curve")
def qpcr_standard_curve(request: StandardCurveRequest) -> Dict[str, Any]:
    points = _standard_curve_points(request)
    if len(points) < 2:
        raise HTTPException(status_code=400, detail="At least two standard points with positive quantity are required")
    x = [math.log(p.quantity, request.log_base) for p in points]
    y = [p.cq for p in points]
    fit = _linear_fit(x, y)
    slope = fit["slope"]
    efficiency = ((request.log_base ** (-1.0 / slope)) - 1.0) * 100.0 if slope else 0.0
    qc_flags: List[str] = []
    if not (-3.9 <= slope <= -2.9):
        qc_flags.append("MIQE slope outside common -3.9 to -2.9 range")
    if not (90 <= efficiency <= 110):
        qc_flags.append("MIQE efficiency outside 90-110% range")
    else:
        qc_flags.append("MIQE efficiency within 90-110% range")
    if fit["r_squared"] < 0.98:
        qc_flags.append("R² below 0.98")
    return _json_clean(
        {
            **fit,
            "efficiency_percent": efficiency,
            "efficiency": efficiency,
            "gene": request.gene,
            "unit": request.unit,
            "n_points": len(points),
            "x_log_quantity": x,
            "concentrations": [p.quantity for p in points],
            "cq_values": y,
            "qc_flags": qc_flags,
            "plotly_json": {
                "data": [
                    {"type": "scatter", "mode": "markers", "x": x, "y": y, "name": "Standards"},
                    {
                        "type": "scatter",
                        "mode": "lines",
                        "x": x,
                        "y": [fit["slope"] * xi + fit["intercept"] for xi in x],
                        "name": "Fit",
                    },
                ],
                "layout": {"title": "qPCR Standard Curve", "xaxis": {"title": "log10 quantity"}, "yaxis": {"title": "Cq"}},
            },
        }
    )


@router.post("/analysis/qpcr/quantify")
def qpcr_quantify(request: QuantifyRequest) -> Dict[str, Any]:
    if request.std_concentrations and request.std_cq_values:
        curve = qpcr_standard_curve(
            StandardCurveRequest(
                concentrations=request.std_concentrations,
                cq_values=request.std_cq_values,
                unit=request.unit,
                log_base=request.log_base,
            )
        )
        sample_cq_values = [float(v) for v in (request.sample_cq_values or request.cq_values) if math.isfinite(float(v))]
        if not sample_cq_values:
            raise HTTPException(status_code=400, detail="At least one finite sample Cq value is required")
        sample_ids = _required_ids(request.sample_ids, len(sample_cq_values), "sample_ids")
        slope = float(curve["slope"])
        intercept = float(curve["intercept"])
        min_std = min(request.std_concentrations)
        max_std = max(request.std_concentrations)
        sample_rows = []
        quantities = []
        for sample_id, cq in zip(sample_ids, sample_cq_values):
            quantity = request.log_base ** ((cq - intercept) / slope)
            quantities.append(quantity)
            within_range = min_std <= quantity <= max_std
            sample_rows.append(
                {
                    "sample_id": sample_id,
                    "cq_value": cq,
                    "quantity": quantity,
                    "quantity_formatted": f"{quantity:.4g} {request.unit}",
                    "unit": request.unit,
                    "within_range": within_range,
                    "extrapolated": not within_range,
                }
            )
        return _json_clean(
            {
                "standard_curve": curve,
                "quantities": sample_rows,
                "mean_quantity": _mean(quantities),
                "sd_quantity": _std(quantities),
                "replicate_cv_percent": _cv_percent(quantities),
                "summary": f"Quantified {len(sample_rows)} qPCR sample(s) using {len(request.std_concentrations)} standards.",
                "plotly_json": {
                    "data": [
                        curve["plotly_json"]["data"][0],
                        curve["plotly_json"]["data"][1],
                        {
                            "type": "scatter",
                            "mode": "markers+text",
                            "x": [math.log(q, request.log_base) for q in quantities if q > 0],
                            "y": sample_cq_values,
                            "text": [row["sample_id"] for row in sample_rows],
                            "name": "Samples",
                        },
                    ],
                    "layout": {"title": "qPCR Absolute Quantification", "xaxis": {"title": "log10 quantity"}, "yaxis": {"title": "Cq"}},
                },
            }
        )

    cq_values = [float(v) for v in request.cq_values if math.isfinite(float(v))]
    if not cq_values:
        raise HTTPException(status_code=400, detail="At least one finite Cq value is required")
    if request.slope is None or request.intercept is None:
        raise HTTPException(status_code=400, detail="slope and intercept are required unless standard curve arrays are provided")
    if request.slope == 0:
        raise HTTPException(status_code=400, detail="Slope cannot be zero")
    quantities = [request.log_base ** ((cq - request.intercept) / request.slope) for cq in cq_values]
    return _json_clean(
        {
            "cq_values": cq_values,
            "quantities": quantities,
            "mean_cq": _mean(cq_values),
            "sd_cq": _std(cq_values),
            "replicate_cv_percent": _cv_percent(quantities),
            "mean_quantity": _mean(quantities),
            "sd_quantity": _std(quantities),
            "qc_flags": ["replicate CV above 20%"] if _cv_percent(quantities) > 20 else [],
        }
    )


def _cq_group(rows: List[CqRow]) -> Dict[Tuple[str, str, str], List[float]]:
    grouped: Dict[Tuple[str, str, str], List[float]] = {}
    for row in rows:
        if math.isfinite(row.cq):
            grouped.setdefault((row.sample, row.group, row.gene), []).append(row.cq)
    return grouped


def _sample_delta_cq(rows: List[CqRow], reference_genes: List[str], target_genes: List[str]) -> List[Dict[str, Any]]:
    grouped = _cq_group(rows)
    by_sample_group: Dict[Tuple[str, str], Dict[str, float]] = {}
    for (sample, group, gene), values in grouped.items():
        by_sample_group.setdefault((sample, group), {})[gene] = _mean(values)
    results: List[Dict[str, Any]] = []
    for (sample, group), gene_means in by_sample_group.items():
        ref_values = [gene_means[g] for g in reference_genes if g in gene_means]
        if not ref_values:
            continue
        ref_mean = _mean(ref_values)
        for target in target_genes:
            if target not in gene_means:
                continue
            delta = gene_means[target] - ref_mean
            results.append(
                {
                    "sample": sample,
                    "group": group,
                    "target_gene": target,
                    "reference_cq_mean": ref_mean,
                    "target_cq_mean": gene_means[target],
                    "delta_cq": delta,
                    "relative_expression": 2 ** (-delta),
                }
            )
    return results


@router.post("/analysis/qpcr/delta-cq")
def qpcr_delta_cq(request: DeltaCqRequest) -> Dict[str, Any]:
    results = _sample_delta_cq(request.rows, request.reference_genes, request.target_genes)
    return {"results": _json_clean(results), "n_results": len(results)}


@router.post("/analysis/qpcr/delta-delta-cq")
def qpcr_delta_delta_cq(request: DeltaDeltaCqRequest) -> Dict[str, Any]:
    delta_rows = _sample_delta_cq(request.rows, request.reference_genes, request.target_genes)
    controls: Dict[str, List[float]] = {}
    for row in delta_rows:
        if row["group"] == request.control_group:
            controls.setdefault(row["target_gene"], []).append(row["delta_cq"])
    control_means = {target: _mean(vals) for target, vals in controls.items()}
    results = []
    for row in delta_rows:
        baseline = control_means.get(row["target_gene"])
        if baseline is None:
            continue
        dd = row["delta_cq"] - baseline
        results.append({**row, "control_delta_cq_mean": baseline, "delta_delta_cq": dd, "fold_change": 2 ** (-dd)})
    return {"results": _json_clean(results), "control_group": request.control_group, "control_means": _json_clean(control_means)}


@router.post("/analysis/qpcr/anova-dunnett")
def qpcr_anova_dunnett(request: AnovaDunnettRequest) -> Dict[str, Any]:
    if isinstance(request.groups, dict):
        named_groups = {}
        for raw_name, vals in request.groups.items():
            name = str(raw_name).strip()
            if not name:
                raise HTTPException(status_code=400, detail="Each group mapping key must be a non-empty real group name")
            if not isinstance(vals, list):
                raise HTTPException(status_code=400, detail=f"Group {name} values must be an array")
            named_groups[name] = [float(v) for v in vals if _finite_float(v) is not None]
    elif isinstance(request.groups, list):
        names = [str(name).strip() for name in request.group_names]
        if len(names) != len(request.groups) or any(not name for name in names):
            raise HTTPException(status_code=400, detail="group_names must provide one non-empty real name for each group array")
        named_groups = {}
        for name, vals in zip(names, request.groups):
            if not isinstance(vals, list):
                raise HTTPException(status_code=400, detail=f"Group {name} values must be an array")
            named_groups[name] = [float(v) for v in vals if _finite_float(v) is not None]
    else:
        raise HTTPException(status_code=400, detail="groups must be a mapping or array of arrays")
    valid = {k: vals for k, vals in named_groups.items() if len(vals) >= 2}
    control_group = request.control_group.strip()
    if not control_group:
        raise HTTPException(status_code=400, detail="control_group must be provided explicitly")
    if control_group not in valid:
        raise HTTPException(status_code=400, detail="Control group must have at least two values")
    if len(valid) < 2:
        raise HTTPException(status_code=400, detail="At least two groups with two finite values are required")
    f_stat, p_value = stats.f_oneway(*valid.values())
    control = valid[control_group]
    comparisons = []
    for name, vals in valid.items():
        if name == control_group:
            continue
        t_stat, p = stats.ttest_ind(vals, control, equal_var=False)
        difference = _mean(vals) - _mean(control)
        comparisons.append(
            {
                "group": name,
                "control_group": control_group,
                "mean_delta": difference,
                "difference": difference,
                "diff": difference,
                "t_statistic": t_stat,
                "p_value": p,
                "significant": p < request.alpha,
            }
        )
    return _json_clean(
        {
            "f_statistic": f_stat,
            "p_value": p_value,
            "anova_f": f_stat,
            "anova_p": p_value,
            "alpha": request.alpha,
            "control_group": control_group,
            "comparisons": comparisons,
            "dunnett_results": comparisons,
            "summary": f"One-way ANOVA across {len(valid)} groups; Dunnett-style comparisons versus {control_group}.",
        }
    )


QPCR_WELL_ALIASES = ("well", "well position", "well_position", "position", "well pos", "well number", "wellnumber")
QPCR_SAMPLE_ALIASES = ("sample name", "sample_name", "sample", "sample id", "sample_id", "specimen", "name")
QPCR_TARGET_ALIASES = ("target name", "target_name", "target", "detector name", "detector", "assay", "gene")
QPCR_TASK_ALIASES = ("task", "type", "sample type", "detector task", "well type")
QPCR_CT_ALIASES = ("ct", "cq", "cт", "ct mean", "cq mean", "ct value", "cq value", "cycle threshold", "threshold cycle", "crt")
QPCR_QUANTITY_ALIASES = ("quantity", "qty", "copies", "copy_number", "copy number", "starting quantity", "starting quantity sq", "sq", "quantity mean", "standard quantity", "amount")
QPCR_GROUP_ALIASES = ("group", "condition", "treatment", "biogroup", "biological group")
QPCR_UNDETERMINED_CT_VALUES = {"undetermined", "undet", "no ct", "no cq", "n/a", "na", "not determined", "not detected", "ntc"}


def _normalize_header_name(value: Any) -> str:
    text = str(value).replace("\ufeff", "").strip().lower()
    text = re.sub(r"[\[\]{}()/:;]+", " ", text)
    text = re.sub(r"[^a-z0-9µμ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _row_lookup(row: Dict[str, Any]) -> Dict[str, Any]:
    lookup: Dict[str, Any] = {}
    for key, value in row.items():
        if key is None:
            continue
        exact = str(key).replace("\ufeff", "").strip().lower()
        if exact:
            lookup.setdefault(exact, value)
        normalized = _normalize_header_name(key)
        if normalized:
            lookup.setdefault(normalized, value)
    return lookup


def _alias_keys(alias: str) -> Tuple[str, str]:
    return str(alias).strip().lower(), _normalize_header_name(alias)


def _has_header_alias(headers: Iterable[str], aliases: Iterable[str]) -> bool:
    header_keys = set()
    for header in headers:
        if header is None:
            continue
        header_keys.update(_alias_keys(str(header)))
    alias_keys = set()
    for alias in aliases:
        alias_keys.update(_alias_keys(str(alias)))
    return bool(header_keys & alias_keys)


def _first_present_value(mapping: Dict[str, Any], aliases: Iterable[str]) -> Optional[Any]:
    for alias in aliases:
        for key in _alias_keys(str(alias)):
            if key in mapping:
                return mapping[key]
    return None


def _first_finite(mapping: Dict[str, Any], aliases: Iterable[str]) -> Optional[float]:
    for alias in aliases:
        for key in _alias_keys(str(alias)):
            if key not in mapping:
                continue
            finite = _finite_float(mapping[key])
            if finite is not None:
                return finite
    return None


def _normalize_qpcr_well_label(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    if not text:
        return None
    match = re.fullmatch(r"([A-H])0?([1-9]|1[0-2])", text)
    if match:
        return f"{match.group(1)}{int(match.group(2))}"
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        well_number = int(float(text))
        if 1 <= well_number <= 96:
            row_idx = (well_number - 1) // 12
            col_idx = (well_number - 1) % 12
            return f"{chr(ord('A') + row_idx)}{col_idx + 1}"
    return text


def _ct_value_or_status(mapping: Dict[str, Any], row_idx: int, context: str) -> Tuple[Optional[float], Optional[str]]:
    saw_undetermined = False
    saw_ct_column = False
    for alias in QPCR_CT_ALIASES:
        for key in _alias_keys(str(alias)):
            if key not in mapping:
                continue
            saw_ct_column = True
            value = mapping[key]
            finite = _finite_float(value)
            if finite is not None:
                return finite, None
            text = str(value).strip().lower() if value is not None else ""
            if not text:
                continue
            if text in QPCR_UNDETERMINED_CT_VALUES:
                saw_undetermined = True
    if saw_undetermined:
        return None, "undetermined"
    problem = "present but non-finite" if saw_ct_column else "missing required"
    raise HTTPException(status_code=400, detail=f"{context} row {row_idx} {problem} Ct/Cq value")


def _parse_qpcr_csv(text: str) -> Tuple[List[Dict[str, Any]], str]:
    csv_rows = [tuple(row) for row in csv.reader(io.StringIO(text))]
    wells = _extract_qpcr_wells_from_row_tables([("CSV", csv_rows)], "qPCR CSV")
    has_plate_positions = any(well.get("well_position") for well in wells)
    instrument_format = "QuantStudio/StepOnePlus CSV" if has_plate_positions else "qPCR Cq table CSV"
    return wells, instrument_format


def _build_qpcr_upload_response(wells: List[Dict[str, Any]], filename: Optional[str], import_engine: str, instrument_format: str) -> Dict[str, Any]:
    if not wells:
        raise HTTPException(status_code=400, detail="No qPCR result rows found. Export with Well/Sample/Target/Ct columns.")
    targets = sorted({w["target_name"] for w in wells})
    samples = sorted({w["sample_name"] for w in wells})
    ct_rows = [w for w in wells if w.get("ct") is not None]

    standard_curves: Dict[str, Dict[str, Any]] = {}
    standard_curve_stats_by_target: Dict[str, Dict[str, Any]] = {}
    quantities: List[Dict[str, Any]] = []
    standard_curve_traces: List[Dict[str, Any]] = []
    curve_palette = ["#38bdf8", "#a78bfa", "#34d399", "#fbbf24", "#fb7185", "#60a5fa"]
    experimental_palette = ["#fb923c", "#f472b6", "#22c55e", "#facc15", "#f87171", "#818cf8"]
    for target_idx, target in enumerate(targets):
        target_color = curve_palette[target_idx % len(curve_palette)]
        experimental_color = experimental_palette[target_idx % len(experimental_palette)]
        std_rows = sorted(
            [
                w
                for w in ct_rows
                if w.get("target_name") == target
                and str(w.get("task") or "").upper() == "STANDARD"
                and _finite_float(w.get("quantity")) is not None
                and float(w["quantity"]) > 0
            ],
            key=lambda w: (math.log(float(w["quantity"]), 10.0), str(w.get("well_position") or "")),
        )
        if len(std_rows) >= 2:
            x = [math.log(float(w["quantity"]), 10.0) for w in std_rows]
            y = [float(w["ct"]) for w in std_rows]
            fit = _linear_fit(x, y)
            slope = fit["slope"]
            efficiency = ((10.0 ** (-1.0 / slope)) - 1.0) * 100.0 if slope else 0.0
            flags: List[str] = []
            if not (-3.9 <= slope <= -2.9):
                flags.append("slope outside common -3.9 to -2.9 range")
            if not (90 <= efficiency <= 110):
                flags.append("efficiency outside 90-110% range")
            if fit["r_squared"] < 0.98:
                flags.append("R² below 0.98")
            stats_payload = {
                **fit,
                "target_name": target,
                "efficiency_percent": efficiency,
                "efficiency": efficiency,
                "n_points": len(std_rows),
                "quantities": [float(w["quantity"]) for w in std_rows],
                "ct_values": y,
                "log10_quantities": x,
                "is_valid": not flags,
                "flags": flags,
            }
            standard_curves[target] = stats_payload
            standard_curve_stats_by_target[target] = stats_payload

            standard_hover = [
                f"{w.get('well_position')} {w.get('sample_name')}<br>{target}<br>Quantity {float(w['quantity']):.6g}<br>Cq {float(w['ct']):.3f}"
                for w in std_rows
            ]
            standard_curve_traces.append(
                {
                    "type": "scatter",
                    "mode": "markers",
                    "x": x,
                    "y": y,
                    "text": [f"{w.get('well_position')} {w.get('sample_name')}" for w in std_rows],
                    "customdata": [float(w["quantity"]) for w in std_rows],
                    "hovertemplate": f"%{{text}}<br>{target}<br>Quantity %{{customdata:.6g}}<br>log10 quantity %{{x:.3f}}<br>Cq %{{y:.3f}}<extra></extra>",
                    "name": f"{target} standards",
                    "legendgroup": target,
                    "marker": {"size": 11, "symbol": "circle", "color": target_color, "opacity": 0.92, "line": {"width": 1.5, "color": "#e0f2fe"}},
                }
            )
            if len(set(x)) > 1:
                line_x = [min(x) + (max(x) - min(x)) * idx / 99 for idx in range(100)]
            else:
                line_x = x
            standard_curve_traces.append(
                {
                    "type": "scatter",
                    "mode": "lines",
                    "x": line_x,
                    "y": [fit["slope"] * xi + fit["intercept"] for xi in line_x],
                    "name": f"{target} fit",
                    "legendgroup": target,
                    "line": {"width": 3.5, "color": target_color, "dash": "solid"},
                    "hovertemplate": f"{target} fit<br>log10 quantity %{{x:.3f}}<br>Cq %{{y:.3f}}<extra></extra>",
                }
            )

            experimental_x: List[float] = []
            experimental_y: List[float] = []
            experimental_text: List[str] = []
            experimental_hover: List[str] = []
            for w in ct_rows:
                task_name = str(w.get("task") or "").upper()
                if w.get("target_name") != target or task_name in {"STANDARD", "NTC"}:
                    continue
                estimated = 10.0 ** ((float(w["ct"]) - fit["intercept"]) / fit["slope"]) if fit["slope"] else None
                log_estimated = math.log(estimated, 10.0) if estimated is not None and estimated > 0 else None
                quantities.append(
                    {
                        "sample_name": w.get("sample_name"),
                        "target_name": target,
                        "well_position": w.get("well_position"),
                        "task": w.get("task"),
                        "ct": w.get("ct"),
                        "estimated_quantity": estimated,
                        "log10_estimated_quantity": log_estimated,
                    }
                )
                if log_estimated is not None:
                    experimental_x.append(log_estimated)
                    experimental_y.append(float(w["ct"]))
                    label = f"{w.get('well_position')} {w.get('sample_name')}"
                    experimental_text.append(label)
                    experimental_hover.append(f"{label}<br>{target}<br>Estimated quantity {estimated:.6g}<br>Cq {float(w['ct']):.3f}")
            if experimental_x:
                standard_curve_traces.append(
                    {
                        "type": "scatter",
                        "mode": "markers",
                        "x": experimental_x,
                        "y": experimental_y,
                        "text": experimental_text,
                        "hovertext": experimental_hover,
                        "hoverinfo": "text",
                        "name": f"{target} experimentals",
                        "legendgroup": target,
                        "marker": {"size": 10, "symbol": "diamond-open", "color": experimental_color, "opacity": 0.95, "line": {"width": 2, "color": "#fed7aa"}},
                    }
                )

    def _plate_row_col(pos_value: Any) -> Optional[Tuple[int, int]]:
        if pos_value is None:
            return None
        pos = str(pos_value).upper().strip()
        if len(pos) < 2 or pos[0] not in "ABCDEFGH":
            return None
        try:
            col = int(pos[1:]) - 1
        except ValueError:
            return None
        row = ord(pos[0]) - ord("A")
        if 0 <= row < 8 and 0 <= col < 12:
            return row, col
        return None

    first_curve_target = next((target for target in targets if target in standard_curves), None)
    first_heatmap_target = first_curve_target or (targets[0] if targets else None)
    heatmap_targets = ([first_heatmap_target] if first_heatmap_target else []) + [target for target in targets if target != first_heatmap_target]
    heatmap_traces: List[Dict[str, Any]] = []
    heatmap_buttons: List[Dict[str, Any]] = []
    for target_idx, target in enumerate(heatmap_targets):
        heatmap_z = [[None for _ in range(12)] for _ in range(8)]
        for w in ct_rows:
            if w.get("target_name") != target:
                continue
            row_col = _plate_row_col(w.get("well_position"))
            if row_col is None:
                continue
            row, col = row_col
            heatmap_z[row][col] = w["ct"]
        heatmap_traces.append(
            {
                "type": "heatmap",
                "z": heatmap_z,
                "x": list(range(1, 13)),
                "y": list("ABCDEFGH"),
                "name": f"{target} Ct",
                "visible": target_idx == 0,
                "colorscale": "Turbo",
                "colorbar": {"title": "Cq/Ct"},
                "hovertemplate": f"{target}<br>Well %{{y}}%{{x}}<br>Cq %{{z:.3f}}<extra></extra>",
            }
        )
        heatmap_buttons.append(
            {
                "label": str(target),
                "method": "update",
                "args": [
                    {"visible": [idx == target_idx for idx in range(len(heatmap_targets))]},
                    {"title": f"qPCR Plate Ct Heatmap — {target}"},
                ],
            }
        )

    replicate_qc: List[Dict[str, Any]] = []
    replicate_groups: Dict[Tuple[Any, Any, Any, Any], List[Dict[str, Any]]] = {}
    for w in ct_rows:
        key = (w.get("sample_name"), w.get("target_name"), w.get("task"), w.get("quantity"))
        replicate_groups.setdefault(key, []).append(w)
    for (sample_name, target_name, task, quantity), rows in sorted(replicate_groups.items(), key=lambda item: tuple(str(part) for part in item[0])):
        values = [float(row["ct"]) for row in rows]
        if len(values) < 2:
            continue
        replicate_qc.append(
            {
                "sample_name": sample_name,
                "target_name": target_name,
                "task": task,
                "quantity": quantity,
                "n": len(values),
                "ct_mean": _mean(values),
                "ct_sd": _std(values),
                "ct_cv_percent": _cv_percent(values),
                "ct_range": max(values) - min(values),
                "well_positions": [row.get("well_position") for row in rows],
            }
        )

    ntc_rows = [w for w in wells if "NTC" in str(w.get("task") or "").upper() or "NTC" in str(w.get("sample_name") or "").upper()]
    ntc_qc = []
    if ntc_rows:
        amplified = [w for w in ntc_rows if w.get("ct") is not None]
        ntc_qc.append(
            {
                "n": len(ntc_rows),
                "amplified_count": len(amplified),
                "min_ct": min([float(w["ct"]) for w in amplified]) if amplified else None,
                "status": "review_amplification" if amplified else "no_threshold_crossing",
            }
        )

    flag_counts: Dict[str, int] = {}
    for w in wells:
        status = w.get("ct_status")
        if status:
            flag_counts[str(status)] = flag_counts.get(str(status), 0) + 1

    amplification_traces: List[Dict[str, Any]] = []
    for w in wells:
        curve = w.get("amplification_curve")
        if not isinstance(curve, dict):
            continue
        cycles = curve.get("cycle")
        delta_rn = curve.get("delta_rn")
        if not cycles or not delta_rn:
            continue
        amplification_traces.append(
            {
                "type": "scatter",
                "mode": "lines",
                "x": cycles,
                "y": delta_rn,
                "name": f"{w.get('well_position')} {w.get('target_name')}",
                "legendgroup": str(w.get("target_name")),
                "opacity": 0.55,
                "line": {"width": 1},
            }
        )

    standard_curve_stats = next(iter(standard_curve_stats_by_target.values()), None)
    return _json_clean(
        {
            "filename": filename,
            "import_engine": import_engine,
            "instrument_format": instrument_format,
            "n_wells": len(wells),
            "targets": targets,
            "samples": samples,
            "wells": wells,
            "results_plotly_json": {
                "data": heatmap_traces,
                "layout": {
                    "title": f"qPCR Plate Ct Heatmap — {first_heatmap_target}" if first_heatmap_target else "qPCR Plate Ct Heatmap",
                    "xaxis": {"title": "Plate column", "dtick": 1},
                    "yaxis": {"title": "Plate row", "autorange": "reversed"},
                    "updatemenus": [{"buttons": heatmap_buttons, "direction": "down", "x": 0.0, "xanchor": "left", "y": 1.16, "yanchor": "top"}] if heatmap_buttons else [],
                    "margin": {"l": 50, "r": 80, "t": 80, "b": 50},
                },
            },
            "amplification_plotly_json": {
                "data": amplification_traces,
                "layout": {"title": "qPCR Amplification Curves", "xaxis": {"title": "Cycle"}, "yaxis": {"title": "ΔRn"}, "hovermode": "closest"},
            }
            if amplification_traces
            else None,
            "standard_curve_plotly_json": {
                "data": standard_curve_traces,
                "layout": {
                    "title": "qPCR Standard Curve + Experimentals",
                    "plot_bgcolor": "rgba(15, 23, 42, 0.35)",
                    "paper_bgcolor": "rgba(0, 0, 0, 0)",
                    "font": {"color": "#e5e7eb"},
                    "xaxis": {"title": "log10 quantity", "showgrid": True, "gridcolor": "rgba(148, 163, 184, 0.22)", "zeroline": False},
                    "yaxis": {"title": "Cq/Ct", "autorange": "reversed", "showgrid": True, "gridcolor": "rgba(148, 163, 184, 0.22)", "zeroline": False},
                    "hovermode": "closest",
                    "legend": {"orientation": "h", "y": -0.25},
                    "margin": {"l": 70, "r": 35, "t": 75, "b": 95},
                },
            }
            if standard_curve_traces
            else None,
            "standard_curve_stats": standard_curve_stats,
            "standard_curve_stats_by_target": standard_curve_stats_by_target,
            "assay_summary": {"standard_curves": standard_curves, "quantities": quantities, "replicate_qc": replicate_qc, "ntc_qc": ntc_qc, "spike_recovery": [], "flag_counts": flag_counts},
        }
    )

async def _persist_qpcr_response_if_requested(response: Dict[str, Any], *, persist: bool, data: bytes, file: UploadFile) -> Dict[str, Any]:
    if persist:
        try:
            durable_ids = await persist_qpcr_import_response(
                response,
                source_bytes=data,
                filename=file.filename,
                content_type=file.content_type,
            )
        except _DB_DEGRADED_EXCEPTIONS as exc:
            _raise_db_service_degraded(exc)
        response.update(durable_ids)
    return response


async def _qpcr_upload_common(file: UploadFile, persist: bool = False) -> Dict[str, Any]:
    data = await file.read()
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    wells, instrument_format = _parse_qpcr_csv(text)
    response = _build_qpcr_upload_response(wells, file.filename, "csv", instrument_format)
    return await _persist_qpcr_response_if_requested(response, persist=persist, data=data, file=file)


def _row_to_qpcr_well(row: Dict[str, Any], idx: int, context: str = "qPCR workbook") -> Dict[str, Any]:
    lower = _row_lookup(row)
    has_well_column = _has_header_alias(lower.keys(), QPCR_WELL_ALIASES)
    well = _normalize_qpcr_well_label(_require_text(lower, QPCR_WELL_ALIASES, "Well", idx, context)) if has_well_column else None
    sample = _require_text(lower, QPCR_SAMPLE_ALIASES, "Sample", idx, context)
    target = _require_text(lower, QPCR_TARGET_ALIASES, "Target", idx, context)
    task = _first_text(lower, QPCR_TASK_ALIASES)
    ct, ct_status = _ct_value_or_status(lower, idx, context)
    qty = _first_finite(lower, QPCR_QUANTITY_ALIASES)
    group = _first_text(lower, QPCR_GROUP_ALIASES)
    parsed = {"well_position": well, "sample_name": sample, "target_name": target, "task": task, "ct": ct, "quantity": qty, "group": group}
    if ct_status is not None:
        parsed["ct_status"] = ct_status
    return parsed


def _extract_qpcr_wells_from_row_tables(sheet_tables: Iterable[Tuple[str, List[Tuple[Any, ...]]]], context: str) -> List[Dict[str, Any]]:
    wells: List[Dict[str, Any]] = []
    for _sheet_name, rows in sheet_tables:
        header_idx: Optional[int] = None
        headers: List[str] = []
        for idx, values in enumerate(rows):
            candidate = [str(v).strip() if v is not None else "" for v in values]
            lower = {_normalize_header_name(v) for v in candidate if v}
            has_ct = _has_header_alias(lower, QPCR_CT_ALIASES)
            has_sample_target = _has_header_alias(lower, QPCR_SAMPLE_ALIASES) and _has_header_alias(lower, QPCR_TARGET_ALIASES)
            has_plate_shape = _has_header_alias(lower, QPCR_WELL_ALIASES)
            if has_ct and (has_sample_target or has_plate_shape):
                header_idx = idx
                headers = candidate
                break
        if header_idx is None:
            continue
        for row_idx, values in enumerate(rows[header_idx + 1 :], start=1):
            row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]}
            if not any(value not in (None, "") for value in row.values()):
                continue
            lower_row = _row_lookup(row)
            has_sample_or_target = _first_text(lower_row, QPCR_SAMPLE_ALIASES) is not None or _first_text(lower_row, QPCR_TARGET_ALIASES) is not None
            has_ct_value = any(_finite_float(_first_present_value(lower_row, [alias])) is not None for alias in QPCR_CT_ALIASES)
            if not has_sample_or_target and not has_ct_value:
                continue
            wells.append(_row_to_qpcr_well(row, row_idx, context))
    return wells


def _parse_qpcr_xls(data: bytes) -> Tuple[List[Dict[str, Any]], str, str]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency is required by pyproject/runtime image
        raise HTTPException(status_code=500, detail="xlrd is not installed in the BMS API environment for legacy .xls parsing") from exc
    try:
        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Unable to parse legacy qPCR XLS workbook with xlrd: {exc}") from exc
    sheet_names = list(workbook.sheet_names())
    sheet_tables: List[Tuple[str, List[Tuple[Any, ...]]]] = []
    for sheet in workbook.sheets():
        sheet_tables.append((sheet.name, [tuple(sheet.row_values(row_idx)) for row_idx in range(sheet.nrows)]))
    wells = _extract_qpcr_wells_from_row_tables(sheet_tables, "qPCR legacy XLS")
    instrument_format = "QuantStudio/StepOnePlus legacy XLS" if any(name.lower() in {"results", "amplification data", "melt region derivative data", "sample setup"} for name in sheet_names) else "generic qPCR legacy XLS"
    return wells, "xlrd", instrument_format


def _parse_qpcr_xlsx(data: bytes) -> Tuple[List[Dict[str, Any]], str, str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is now required by pyproject
        raise HTTPException(status_code=500, detail="openpyxl is not installed in the BMS API environment") from exc

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    sheet_tables = [(sheet.title, [tuple(row) for row in sheet.iter_rows(values_only=True)]) for sheet in workbook.worksheets]
    wells = _extract_qpcr_wells_from_row_tables(sheet_tables, "qPCR workbook")
    instrument_format = "QuantStudio/StepOnePlus Excel" if any(name.lower() in {"results", "amplification data", "melt region derivative data", "sample setup"} for name in sheet_names) else "generic qPCR Excel"
    return wells, "openpyxl", instrument_format


def _parse_qpcr_excel(data: bytes, filename: Optional[str] = None) -> Tuple[List[Dict[str, Any]], str, str]:
    name = (filename or "").lower()
    if name.endswith(".xls") and not name.endswith((".xlsx", ".xlsm")):
        return _parse_qpcr_xls(data)
    try:
        return _parse_qpcr_xlsx(data)
    except HTTPException:
        raise
    except Exception as exc:
        if name.endswith(".xls") and not name.endswith((".xlsx", ".xlsm")):
            return _parse_qpcr_xls(data)
        raise exc


def _xml_children(element: ET.Element, tag_name: str) -> List[ET.Element]:
    return [child for child in list(element) if child.tag.rsplit("}", 1)[-1] == tag_name]


def _xml_first(element: Optional[ET.Element], tag_name: str) -> Optional[ET.Element]:
    if element is None:
        return None
    for child in list(element):
        if child.tag.rsplit("}", 1)[-1] == tag_name:
            return child
    return None


def _xml_text(element: Optional[ET.Element], tag_name: str) -> Optional[str]:
    child = _xml_first(element, tag_name)
    text = (child.text or "").strip() if child is not None else ""
    return text if text else None


def _safe_int(value: Optional[str], default: int) -> int:
    try:
        return int(float(str(value)))
    except (TypeError, ValueError):
        return default


def _coerce_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y", "on"}:
        return True
    if text in {"false", "0", "no", "n", "off"}:
        return False
    return default


def _eds_well_position(index: int, columns: int) -> str:
    columns = max(1, columns)
    row_idx = index // columns
    col_idx = index % columns
    if 0 <= row_idx < 26:
        return f"{chr(ord('A') + row_idx)}{col_idx + 1}"
    return str(index + 1)


def _parse_eds_dye_list(text: Optional[str]) -> List[str]:
    if not text:
        return []
    cleaned = text.strip().strip("[]")
    return [piece.strip() for piece in cleaned.split(",") if piece.strip()]


def _xml_attr(element: Optional[ET.Element], attr_name: str) -> Optional[str]:
    if element is None:
        return None
    for key, value in element.attrib.items():
        if key.rsplit("}", 1)[-1].lower() == attr_name.lower():
            text = str(value).strip()
            return text if text else None
    return None


def _parse_eds_numeric_array(text: Optional[str]) -> List[float]:
    if not text:
        return []
    cleaned = text.strip().strip("[]")
    values: List[float] = []
    for piece in cleaned.split(","):
        value = _finite_float(piece.strip())
        if value is not None:
            values.append(value)
    return values


def _eds_ct_settings_from_generic(generic: Dict[str, Any]) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Any]]:
    default_threshold = _finite_float(generic.get("threshold"), 0.2) or 0.2
    default_start = _safe_int(generic.get("defaultBaseLineStart") or generic.get("defaultBaselineStart"), 3)
    default_end = _safe_int(generic.get("defaultBaseLineEnd") or generic.get("defaultBaselineEnd"), 15)
    default_auto_baseline = _coerce_bool(generic.get("autoBaseLine") if generic.get("autoBaseLine") is not None else generic.get("autoBaseline"), False)
    default_auto_threshold = _coerce_bool(generic.get("autoThreshold"), False)
    default = {
        "threshold": default_threshold,
        "baseline_start": default_start,
        "baseline_end": default_end,
        "auto_baseline": default_auto_baseline,
        "auto_threshold": default_auto_threshold,
    }
    by_target: Dict[str, Dict[str, Any]] = {}
    settings = generic.get("ctSettingsDetailsDTOS") or generic.get("ctSettingsDetailsDTOs") or generic.get("ctSettingsDetailsDtos") or []
    if isinstance(settings, list):
        for item in settings:
            if not isinstance(item, dict):
                continue
            target = str(item.get("target") or item.get("name") or "").strip()
            if not target:
                continue
            threshold = _finite_float(item.get("threshold"), default_threshold) or default_threshold
            baseline_start = _safe_int(item.get("baselineStart") if item.get("baselineStart") is not None else item.get("defaultBaselineStart"), default_start)
            baseline_end = _safe_int(item.get("baselineEnd") if item.get("baselineEnd") is not None else item.get("defaultBaselineEnd"), default_end)
            by_target[target] = {
                "threshold": threshold,
                "baseline_start": baseline_start,
                "baseline_end": baseline_end,
                "auto_baseline": _coerce_bool(item.get("autoBaseline"), default_auto_baseline),
                "auto_threshold": _coerce_bool(item.get("autoThreshold"), default_auto_threshold),
            }
    return by_target, default


def _eds_jaxb_value(value_item: ET.Element) -> Any:
    for child in list(value_item):
        child_name = child.tag.rsplit("}", 1)[-1]
        text = (child.text or "").strip()
        if not text:
            continue
        if child_name == "StringValue":
            return text
        if child_name == "DoubleValue":
            return _finite_float(text)
        if child_name == "BooleanValue":
            return _coerce_bool(text)
        if child_name in {"IntegerValue", "IntValue", "LongValue"}:
            return _safe_int(text, 0)
        return text
    return None


def _eds_analysis_settings_from_analysis_protocol(xml_bytes: bytes) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Any]]:
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return {}, {}
    out: Dict[str, Dict[str, Any]] = {}
    global_settings: Dict[str, Any] = {}
    for settings in root.iter():
        if settings.tag.rsplit("}", 1)[-1] != "JaxbAnalysisSettings":
            continue
        values: Dict[str, Any] = {}
        for setting_value in _xml_children(settings, "JaxbSettingValue"):
            name = _xml_text(setting_value, "Name")
            value_item = _xml_first(setting_value, "JaxbValueItem")
            if not name or value_item is None:
                continue
            value = _eds_jaxb_value(value_item)
            if value is not None:
                values[name] = value
        object_name = str(values.get("ObjectName") or "").strip()
        target = "" if object_name == "AnalysisProtocol.DEFAULT_SETTINGS" else object_name
        if target:
            threshold = _finite_float(values.get("Threshold"), 0.2) or 0.2
            baseline_start = _safe_int(values.get("BaselineStart"), 3)
            baseline_end = _safe_int(values.get("BaselineStop") if values.get("BaselineStop") is not None else values.get("BaselineEnd"), 15)
            out[target] = {
                "threshold": threshold,
                "baseline_start": baseline_start,
                "baseline_end": baseline_end,
                "auto_baseline": _coerce_bool(values.get("AutoBaseline"), False),
                "auto_threshold": _coerce_bool(values.get("AutoThreshold"), False),
            }
            continue
        if "Threshold" in values:
            global_settings["threshold"] = _finite_float(values.get("Threshold"), 0.2) or 0.2
        if "BaselineStart" in values:
            global_settings["baseline_start"] = _safe_int(values.get("BaselineStart"), 3)
        if "BaselineStop" in values or "BaselineEnd" in values:
            global_settings["baseline_end"] = _safe_int(values.get("BaselineStop") if values.get("BaselineStop") is not None else values.get("BaselineEnd"), 15)
        if "AutoBaseline" in values:
            global_settings["auto_baseline"] = _coerce_bool(values.get("AutoBaseline"), False)
        if "AutoThreshold" in values:
            global_settings["auto_threshold"] = _coerce_bool(values.get("AutoThreshold"), False)
        if "SignalSmoothing" in values:
            global_settings["signal_smoothing"] = _coerce_bool(values.get("SignalSmoothing"), False)
        if "AlgorithmName" in values:
            global_settings["algorithm_name"] = str(values.get("AlgorithmName"))
        if "AutoAnalysis" in values:
            global_settings["auto_analysis"] = _coerce_bool(values.get("AutoAnalysis"), False)
    return out, global_settings


def _eds_ct_settings_from_analysis_protocol(xml_bytes: bytes) -> Dict[str, Dict[str, Any]]:
    return _eds_analysis_settings_from_analysis_protocol(xml_bytes)[0]


def _eds_linear_baseline(normalized: List[float], baseline_start: int, baseline_end: int) -> Optional[Dict[str, Any]]:
    n = len(normalized)
    if n == 0:
        return None
    start_cycle = max(1, min(n, int(baseline_start)))
    end_cycle = max(start_cycle, min(n, int(baseline_end)))
    x = np.arange(start_cycle, end_cycle + 1, dtype=float)
    y = np.asarray(normalized[start_cycle - 1 : end_cycle], dtype=float)
    finite_mask = np.isfinite(y)
    if int(np.count_nonzero(finite_mask)) == 0:
        return None
    x = x[finite_mask]
    y = y[finite_mask]
    if len(x) >= 2:
        slope, intercept = np.polyfit(x, y, 1)
    else:
        slope = 0.0
        intercept = float(y[0])
    cycles = np.arange(1, n + 1, dtype=float)
    baseline = slope * cycles + intercept
    residuals = y - (slope * x + intercept)
    return {
        "baseline": baseline.astype(float).tolist(),
        "baseline_start": start_cycle,
        "baseline_end": end_cycle,
        "baseline_slope": float(slope),
        "baseline_intercept": float(intercept),
        "baseline_residual_sd": _std(residuals.tolist()) if len(residuals) > 1 else 0.0,
    }


def _eds_smooth_delta_rn(delta_rn: List[float], signal_smoothing: bool) -> List[float]:
    if not signal_smoothing:
        return [float(value) for value in delta_rn]
    n = len(delta_rn)
    if n < 5:
        return [float(value) for value in delta_rn]
    window = 5 if n >= 5 else n
    if window % 2 == 0:
        window -= 1
    if window < 5:
        return [float(value) for value in delta_rn]
    try:
        return [float(value) for value in savgol_filter(np.asarray(delta_rn, dtype=float), window, 2, mode="interp")]
    except Exception:
        return [float(value) for value in delta_rn]


def _eds_threshold_crossing(delta_rn: List[float], threshold: float, scan_after_cycle: int, interpolation: str = "linear") -> Optional[float]:
    if not delta_rn:
        return None
    y = np.asarray(delta_rn, dtype=float)
    cycles = np.arange(1, len(y) + 1, dtype=float)
    scan_start = min(len(y) - 1, max(1, int(scan_after_cycle)))
    for idx in range(scan_start, len(y)):
        prev = y[idx - 1]
        cur = y[idx]
        if not (math.isfinite(float(prev)) and math.isfinite(float(cur))):
            continue
        if prev < threshold <= cur and cur != prev:
            linear_estimate = float(cycles[idx - 1] + (threshold - prev) / (cur - prev))
            if interpolation != "cubic":
                return linear_estimate
            local_start = max(0, idx - 2)
            local_end = min(len(y), idx + 3)
            local_x = cycles[local_start:local_end]
            local_y = y[local_start:local_end] - threshold
            degree = min(3, len(local_x) - 1)
            if degree >= 1:
                try:
                    coeff = np.polyfit(local_x, local_y, degree)
                    roots = np.roots(coeff)
                    candidates = [float(root.real) for root in roots if abs(float(root.imag)) < 1e-6 and cycles[idx - 1] - 1e-9 <= float(root.real) <= cycles[idx] + 1e-9]
                    if candidates:
                        return min(candidates, key=lambda candidate: abs(candidate - linear_estimate))
                except Exception:
                    pass
            return linear_estimate
        if idx == scan_start and cur >= threshold:
            return float(cycles[idx])
    return None


def _eds_ct_for_baseline(
    normalized: List[float],
    threshold: float,
    baseline_start: int,
    baseline_end: int,
    signal_smoothing: bool,
) -> Dict[str, Any]:
    baseline_result = _eds_linear_baseline(normalized, baseline_start, baseline_end)
    if baseline_result is None:
        return {"ct": None, "ct_status": "invalid_baseline_window"}
    baseline = baseline_result["baseline"]
    delta_rn = [float(value) - float(base_value) for value, base_value in zip(normalized, baseline)]
    smoothed_delta_rn = _eds_smooth_delta_rn(delta_rn, signal_smoothing)
    interpolation = "cubic" if signal_smoothing else "linear"
    ct_value = _eds_threshold_crossing(smoothed_delta_rn, threshold, int(baseline_result["baseline_end"]), interpolation=interpolation)
    payload: Dict[str, Any] = {
        **{key: value for key, value in baseline_result.items() if key != "baseline"},
        "baseline_rn": _mean(baseline[baseline_result["baseline_start"] - 1 : baseline_result["baseline_end"]]),
        "baseline_method": "linear_regression",
        "threshold_interpolation": interpolation,
        "threshold": threshold,
        "signal_smoothing": signal_smoothing,
        "amplification_curve": {
            "cycle": list(range(1, len(normalized) + 1)),
            "delta_rn": delta_rn,
            "delta_rn_smoothed": smoothed_delta_rn if signal_smoothing else None,
            "rn": normalized,
            "linear_baseline_rn": baseline,
        },
    }
    if ct_value is None:
        payload.update({"ct": None, "ct_status": "no_threshold_crossing"})
    else:
        source_suffix = "cubic_threshold" if interpolation == "cubic" else "linear_threshold"
        payload.update({"ct": ct_value, "ct_source": f"multicomponentdata_linear_baseline_{source_suffix}"})
    return payload


def _calculate_ct_from_eds_curves(
    reporter_curve: List[float],
    passive_curve: Optional[List[float]],
    threshold: float,
    baseline_start: int,
    baseline_end: int,
    auto_baseline: bool = False,
    signal_smoothing: bool = False,
) -> Dict[str, Any]:
    if not reporter_curve:
        return {"ct": None, "ct_status": "missing_reporter_curve"}
    n = len(reporter_curve)
    normalized: List[float] = []
    for idx, reporter_value in enumerate(reporter_curve):
        passive_value = passive_curve[idx] if passive_curve is not None and idx < len(passive_curve) else None
        if passive_value is not None and math.isfinite(passive_value) and passive_value != 0:
            normalized.append(float(reporter_value) / float(passive_value))
        else:
            normalized.append(float(reporter_value))
    configured_end = max(1, min(n, baseline_end))
    selected_end = configured_end
    preliminary_ct: Optional[float] = None
    if auto_baseline and n >= 8:
        preliminary = _eds_ct_for_baseline(normalized, threshold, baseline_start, configured_end, signal_smoothing)
        preliminary_ct = _finite_float(preliminary.get("ct"))
        if preliminary_ct is not None:
            min_end = max(baseline_start + 1, 2)
            max_end = max(min_end, min(n - 3, 30))
            selected_end = max(min_end, min(max_end, int(round(preliminary_ct - 7.0))))
    result = _eds_ct_for_baseline(normalized, threshold, baseline_start, selected_end, signal_smoothing)
    result["auto_baseline"] = bool(auto_baseline)
    result["baseline_end_configured"] = configured_end
    if preliminary_ct is not None:
        result["preliminary_ct"] = preliminary_ct
        result["auto_baseline_offset_cycles"] = 7.0
    return result


def _parse_quantstudio_eds_zip_xml(data: bytes, qslib_error: Optional[str] = None) -> Tuple[List[Dict[str, Any]], Dict[str, Any], List[str]]:
    try:
        archive = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as exc:
        raise HTTPException(status_code=400, detail=f"Unable to parse QuantStudio EDS ZIP container: {exc}") from exc
    with archive:
        names = set(archive.namelist())
        plate_name = next((name for name in names if name.lower().endswith("plate_setup.xml")), None)
        if plate_name is None:
            raise HTTPException(status_code=400, detail="QuantStudio EDS ZIP did not contain plate_setup.xml")
        try:
            plate_root = ET.fromstring(archive.read(plate_name))
        except ET.ParseError as exc:
            raise HTTPException(status_code=400, detail=f"Unable to parse EDS plate_setup.xml: {exc}") from exc

        columns = _safe_int(_xml_text(plate_root, "Columns") or _xml_text(plate_root, "ColumnCount"), 12)
        rows_count = _safe_int(_xml_text(plate_root, "Rows") or _xml_text(plate_root, "RowCount"), 8)
        passive_reference = _xml_text(plate_root, "PassiveReferenceDye")
        sample_by_index: Dict[int, str] = {}
        detector_tasks_by_index: Dict[int, List[Dict[str, Any]]] = {}
        for feature_map in _xml_children(plate_root, "FeatureMap"):
            feature = _xml_first(feature_map, "Feature")
            feature_id = (_xml_text(feature, "Id") or _xml_text(feature, "Name") or "").strip().lower()
            for feature_value in _xml_children(feature_map, "FeatureValue"):
                index = _safe_int(_xml_text(feature_value, "Index"), -1)
                if index < 0:
                    continue
                feature_item = _xml_first(feature_value, "FeatureItem")
                if feature_id == "sample":
                    sample = _xml_first(feature_item, "Sample")
                    sample_name = _xml_text(sample, "Name")
                    if sample_name:
                        sample_by_index[index] = sample_name
                elif feature_id == "detector-task":
                    task_list = _xml_first(feature_item, "DetectorTaskList")
                    tasks: List[Dict[str, Any]] = []
                    for detector_task in _xml_children(task_list, "DetectorTask"):
                        detector = _xml_first(detector_task, "Detector")
                        target_name = _xml_text(detector, "Name")
                        if not target_name:
                            continue
                        tasks.append(
                            {
                                "task": _xml_text(detector_task, "Task"),
                                "quantity": _finite_float(_xml_text(detector_task, "Concentration")),
                                "target_name": target_name,
                                "reporter": _xml_text(detector, "Reporter"),
                                "quencher": _xml_text(detector, "Quencher"),
                            }
                        )
                    if tasks:
                        detector_tasks_by_index[index] = tasks

        generic_name = next((name for name in names if name.lower().endswith("generic_properties.json")), None)
        generic_summary: Dict[str, Any] = {}
        ct_settings_by_target: Dict[str, Dict[str, Any]] = {}
        default_ct_settings: Dict[str, Any] = {"threshold": 0.2, "baseline_start": 3, "baseline_end": 15, "auto_baseline": False, "auto_threshold": False}
        analysis_protocol_settings: Dict[str, Any] = {}
        if generic_name is not None:
            try:
                generic = json.loads(archive.read(generic_name).decode("utf-8"))
                if isinstance(generic, dict):
                    for key in ("assayName", "runBy", "runSoftwareVersion", "exportedSoftwareVersion", "runStartDate", "runEndDate"):
                        if generic.get(key) is not None:
                            generic_summary[key] = generic[key]
                    samples = generic.get("samples")
                    if isinstance(samples, list):
                        generic_summary["sample_metadata_count"] = len(samples)
                    ct_settings_by_target, default_ct_settings = _eds_ct_settings_from_generic(generic)
            except (json.JSONDecodeError, UnicodeDecodeError):
                generic_summary = {}

        analysis_protocol_name = next((name for name in names if name.lower().endswith("analysis_protocol.xml")), None)
        if analysis_protocol_name is not None:
            protocol_settings, analysis_protocol_settings = _eds_analysis_settings_from_analysis_protocol(archive.read(analysis_protocol_name))
            default_ct_settings = {**default_ct_settings, **{key: value for key, value in analysis_protocol_settings.items() if key in {"threshold", "baseline_start", "baseline_end", "auto_baseline", "auto_threshold"}}}
            for target, settings in protocol_settings.items():
                ct_settings_by_target.setdefault(target, settings)

        multicomponent_name = next((name for name in names if name.lower().endswith("multicomponentdata.xml")), None)
        well_count: Optional[int] = None
        cycle_count: Optional[int] = None
        all_dyes: List[str] = []
        dye_by_index: Dict[int, List[str]] = {}
        curves_by_index: Dict[int, Dict[str, List[float]]] = {}
        if multicomponent_name is not None:
            try:
                multicomponent_root = ET.fromstring(archive.read(multicomponent_name))
                well_count = _safe_int(_xml_text(multicomponent_root, "WellCount"), 0) or None
                cycle_count = _safe_int(_xml_text(multicomponent_root, "CycleCount"), 0) or None
                dye_seen: List[str] = []
                for dye_data in _xml_children(multicomponent_root, "DyeData"):
                    well_index = _safe_int(_xml_attr(dye_data, "WellIndex") or _xml_text(dye_data, "WellIndex"), -1)
                    dyes = _parse_eds_dye_list(_xml_text(dye_data, "DyeList"))
                    if well_index >= 0 and dyes:
                        dye_by_index[well_index] = dyes
                    for dye in dyes:
                        if dye not in dye_seen:
                            dye_seen.append(dye)
                all_dyes = dye_seen
                for signal_data in _xml_children(multicomponent_root, "SignalData"):
                    well_index = _safe_int(_xml_attr(signal_data, "WellIndex") or _xml_text(signal_data, "WellIndex"), -1)
                    if well_index < 0:
                        continue
                    dyes = dye_by_index.get(well_index, [])
                    cycle_arrays = [_parse_eds_numeric_array((child.text or "").strip()) for child in _xml_children(signal_data, "CycleData")]
                    curves: Dict[str, List[float]] = {}
                    for idx, values in enumerate(cycle_arrays):
                        if not values:
                            continue
                        dye_name = dyes[idx] if idx < len(dyes) else f"channel_{idx + 1}"
                        curves[dye_name] = values
                    if curves:
                        curves_by_index[well_index] = curves
            except ET.ParseError:
                all_dyes = []
                dye_by_index = {}
                curves_by_index = {}

        wells: List[Dict[str, Any]] = []
        calculated_ct_count = 0
        missing_curve_count = 0
        no_crossing_count = 0
        for index in sorted(set(sample_by_index) & set(detector_tasks_by_index)):
            sample_name = sample_by_index[index]
            curves = curves_by_index.get(index, {})
            for task in detector_tasks_by_index[index]:
                target_name = task["target_name"]
                reporter = task.get("reporter")
                settings = ct_settings_by_target.get(target_name, default_ct_settings)
                threshold = _finite_float(settings.get("threshold"), 0.2) or 0.2
                baseline_start = _safe_int(settings.get("baseline_start"), 3)
                baseline_end = _safe_int(settings.get("baseline_end"), 15)
                auto_baseline = _coerce_bool(settings.get("auto_baseline"), _coerce_bool(default_ct_settings.get("auto_baseline"), False))
                signal_smoothing = _coerce_bool(analysis_protocol_settings.get("signal_smoothing"), False)
                reporter_curve = curves.get(str(reporter)) if reporter else None
                passive_curve = curves.get(str(passive_reference)) if passive_reference else None
                task_name = str(task.get("task") or "").upper()
                well: Dict[str, Any] = {
                    "well_position": _eds_well_position(index, columns),
                    "well_index": index,
                    "sample_name": sample_name,
                    "target_name": target_name,
                    "task": task.get("task"),
                    "ct": None,
                    "quantity": task.get("quantity") if task_name == "STANDARD" else None,
                    "reporter": reporter,
                    "quencher": task.get("quencher"),
                    "passive_reference": passive_reference,
                }
                if reporter_curve:
                    ct_result = _calculate_ct_from_eds_curves(reporter_curve, passive_curve, threshold, baseline_start, baseline_end, auto_baseline=auto_baseline, signal_smoothing=signal_smoothing)
                    well.update(ct_result)
                    if well.get("ct") is not None:
                        calculated_ct_count += 1
                    elif well.get("ct_status") == "no_threshold_crossing":
                        no_crossing_count += 1
                else:
                    well["ct_status"] = "missing_reporter_curve"
                    missing_curve_count += 1
                wells.append(well)

        quant_file_count = sum(1 for name in names if name.lower().endswith(".quant"))
        available_data = sorted(
            item
            for item in [
                "plate_setup.xml" if plate_name else None,
                "multicomponentdata.xml" if multicomponent_name else None,
                "generic_properties.json" if generic_name else None,
                "analysis_protocol.xml" if analysis_protocol_name else None,
                "quant_raw_cycle_files" if quant_file_count else None,
            ]
            if item is not None
        )
        summary: Dict[str, Any] = {
            **generic_summary,
            "well_count": well_count,
            "plate_rows": rows_count,
            "plate_columns": columns,
            "cycle_count": cycle_count,
            "dyes": all_dyes,
            "passive_reference": passive_reference,
            "quant_file_count": quant_file_count,
            "plate_rows_with_sample_and_detector_metadata": len(wells),
            "ct_result_table_detected": False,
            "ct_values_calculated_from_multicomponentdata": calculated_ct_count,
            "ct_values_missing_reporter_curve": missing_curve_count,
            "ct_values_without_threshold_crossing": no_crossing_count,
            "ct_algorithm": "reporter/passive Rn + auto-linear-baseline + Savitzky-Golay smoothing + cubic threshold interpolation"
            if _coerce_bool(analysis_protocol_settings.get("signal_smoothing"), False)
            else "reporter/passive Rn + configured linear baseline + linear threshold interpolation",
            "ct_values_are_authoritative": False,
            "ct_provenance": "computed_from_multicomponentdata_no_scalar_result_table",
            "default_ct_settings": default_ct_settings,
            "analysis_protocol_settings": analysis_protocol_settings,
            "recommendation": "Use the QuantStudio/StepOnePlus Excel Results export as the authoritative Ct/Cq and result Quantity source when available; EDS-derived Ct values are computed from raw curves because this EDS archive does not expose a scalar result table.",
            "note": "Cq values were calculated from real EDS multicomponent amplification curves and per-target threshold/baseline settings; no scalar Cq table was present in the EDS archive.",
        }
        if qslib_error:
            summary["qslib_error"] = qslib_error
        if not wells:
            raise HTTPException(status_code=400, detail="No EDS sample/detector task rows found in plate_setup.xml")
        if calculated_ct_count == 0 and curves_by_index:
            summary["warning"] = "EDS amplification curves were present, but no detector rows crossed configured thresholds"
        return wells, summary, available_data

@router.post("/analysis/qpcr/upload-csv")
async def upload_qpcr_csv(file: UploadFile = File(...), persist: bool = Form(False)) -> Dict[str, Any]:
    return await _qpcr_upload_common(file, persist=persist)


@router.post("/analysis/qpcr/upload-excel")
async def upload_qpcr_excel(file: UploadFile = File(...), persist: bool = Form(False)) -> Dict[str, Any]:
    data = await file.read()
    try:
        wells, import_engine, instrument_format = _parse_qpcr_excel(data, file.filename)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Unable to parse qPCR Excel workbook: {exc}") from exc
    response = _build_qpcr_upload_response(wells, file.filename, import_engine, instrument_format)
    return await _persist_qpcr_response_if_requested(response, persist=persist, data=data, file=file)


@router.post("/analysis/qpcr/upload-eds")
async def upload_qpcr_eds(file: UploadFile = File(...), persist: bool = Form(False)) -> Dict[str, Any]:
    data = await file.read()
    try:
        import qslib
    except ImportError:
        wells, eds_summary, available_data = _parse_quantstudio_eds_zip_xml(data, qslib_error="qslib is not installed in the BMS API environment")
        response = _build_qpcr_upload_response(wells, file.filename, "quantstudio_eds_zip_xml", "QuantStudio EDS ZIP/XML")
        response["eds_summary"] = eds_summary
        response["available_data"] = available_data
        return await _persist_qpcr_response_if_requested(response, persist=persist, data=data, file=file)
    try:
        experiment = qslib.Experiment.from_file(io.BytesIO(data))
    except Exception as exc:
        wells, eds_summary, available_data = _parse_quantstudio_eds_zip_xml(data, qslib_error=str(exc))
        response = _build_qpcr_upload_response(wells, file.filename, "quantstudio_eds_zip_xml", "QuantStudio EDS ZIP/XML")
        response["eds_summary"] = eds_summary
        response["available_data"] = available_data
        return await _persist_qpcr_response_if_requested(response, persist=persist, data=data, file=file)

    wells: List[Dict[str, Any]] = []
    try:
        plate_setup = getattr(experiment, "plate_setup", None) or getattr(experiment, "platesetup", None)
        if plate_setup is not None and hasattr(plate_setup, "to_table"):
            table = plate_setup.to_table()
            for idx, row in enumerate(table if isinstance(table, list) else [], start=1):
                if isinstance(row, dict):
                    wells.append(_row_to_qpcr_well(row, idx))
    except HTTPException:
        wells, eds_summary, available_data = _parse_quantstudio_eds_zip_xml(data, qslib_error="qslib plate table did not expose complete Sample/Target/Ct rows")
        response = _build_qpcr_upload_response(wells, file.filename, "quantstudio_eds_zip_xml", "QuantStudio EDS ZIP/XML")
        response["eds_summary"] = eds_summary
        response["available_data"] = available_data
        return await _persist_qpcr_response_if_requested(response, persist=persist, data=data, file=file)
    except Exception as exc:
        wells, eds_summary, available_data = _parse_quantstudio_eds_zip_xml(data, qslib_error=f"qslib opened the EDS file but BMS could not extract a Cq table: {exc}")
        response = _build_qpcr_upload_response(wells, file.filename, "quantstudio_eds_zip_xml", "QuantStudio EDS ZIP/XML")
        response["eds_summary"] = eds_summary
        response["available_data"] = available_data
        return await _persist_qpcr_response_if_requested(response, persist=persist, data=data, file=file)

    if not wells:
        wells, eds_summary, available_data = _parse_quantstudio_eds_zip_xml(data, qslib_error="qslib opened the EDS file but returned no plate rows")
        response = _build_qpcr_upload_response(wells, file.filename, "quantstudio_eds_zip_xml", "QuantStudio EDS ZIP/XML")
        response["eds_summary"] = eds_summary
        response["available_data"] = available_data
        return await _persist_qpcr_response_if_requested(response, persist=persist, data=data, file=file)

    response = _build_qpcr_upload_response(wells, file.filename, "qslib", "QuantStudio EDS")
    response["eds_summary"] = getattr(experiment, "summary", None)
    response["available_data"] = list(getattr(experiment, "available_data", []) or [])
    return await _persist_qpcr_response_if_requested(response, persist=persist, data=data, file=file)


@router.get("/analysis/qpcr/imports")
async def qpcr_imports(limit: int = Query(50, ge=1, le=200)) -> List[Dict[str, Any]]:
    try:
        return _json_clean(await list_qpcr_imports(limit=limit))
    except _DB_DEGRADED_EXCEPTIONS as exc:
        _raise_db_service_degraded(exc)


@router.get("/datasets")
async def assay_datasets(assay_type: Optional[str] = None, limit: int = Query(100, ge=1, le=500)) -> List[Dict[str, Any]]:
    try:
        return _json_clean(await list_analytical_datasets(assay_type=assay_type, limit=limit))
    except _DB_DEGRADED_EXCEPTIONS as exc:
        _raise_db_service_degraded(exc)


@router.get("/datasets/{dataset_id}")
async def assay_dataset_detail(dataset_id: str) -> Dict[str, Any]:
    try:
        payload = await load_analytical_dataset(dataset_id)
    except _DB_DEGRADED_EXCEPTIONS as exc:
        _raise_db_service_degraded(exc)
    if payload is None:
        raise HTTPException(status_code=404, detail="Analytical dataset not found")
    return _json_clean(payload)


@router.get("/analysis/qpcr/imports/{analytical_import_id}")
async def qpcr_import_detail(analytical_import_id: str) -> Dict[str, Any]:
    try:
        payload = await load_qpcr_import(analytical_import_id)
    except _DB_DEGRADED_EXCEPTIONS as exc:
        _raise_db_service_degraded(exc)
    if payload is None:
        raise HTTPException(status_code=404, detail="qPCR analytical import not found")
    return _json_clean(payload)


class ChromatographyAnalyzeRequest(BaseModel):
    time: List[float]
    signal: List[float]
    baseline_method: str = "rolling_min"
    peak_prominence: float = 0.05
    fit_model: str = "gaussian"


class CalibrationPoint(BaseModel):
    concentration: float
    area: float


class HplcCalibrationRequest(BaseModel):
    points: List[CalibrationPoint]
    analyte_name: str = "Analyte"
    unit: str = "ng/uL"
    force_through_origin: bool = False


class HplcQuantifyRequest(BaseModel):
    area: Optional[float] = None
    slope: Optional[float] = None
    intercept: float = 0.0
    dilution_factor: float = 1.0
    cal_concentrations: List[float] = Field(default_factory=list)
    cal_areas: List[float] = Field(default_factory=list)
    sample_areas: List[float] = Field(default_factory=list)
    sample_ids: List[str] = Field(default_factory=list)
    unit: str = "ng/uL"


class PlasmidIsoformRequest(BaseModel):
    peaks: List[Dict[str, Any]]
    windows: Dict[str, Tuple[float, float]]


@router.post("/analysis/hplc/analyze")
def chromatography_analyze(request: ChromatographyAnalyzeRequest) -> Dict[str, Any]:
    if len(request.time) != len(request.signal) or len(request.time) < 3:
        raise HTTPException(status_code=400, detail="time and signal arrays must be same length with at least 3 points")
    time = np.asarray(request.time, dtype=float)
    signal = np.asarray(request.signal, dtype=float)
    if not np.all(np.isfinite(time)) or not np.all(np.isfinite(signal)):
        raise HTTPException(status_code=400, detail="time and signal must be finite")
    analysis_engine = "scipy.signal"
    engine_package = "scipy"
    if request.baseline_method == "none":
        baseline = np.zeros_like(signal)
    elif request.baseline_method == "linear":
        baseline = np.linspace(signal[0], signal[-1], len(signal))
    elif request.baseline_method.startswith("mocca2"):
        try:
            import mocca2
        except ImportError as exc:  # pragma: no cover - dependency is required by pyproject
            raise HTTPException(status_code=500, detail="mocca2 is not installed in the BMS API environment") from exc
        method = request.baseline_method.removeprefix("mocca2_") or "flatfit"
        if method not in {"flatfit", "arpls", "asls"}:
            method = "flatfit"
        baseline = np.asarray(mocca2.estimate_baseline(signal, method=method), dtype=float)
        analysis_engine = "MOCCA2"
        engine_package = "mocca2"
    else:
        window = max(3, min(len(signal), len(signal) // 20 or 3))
        baseline = np.array([np.min(signal[max(0, i - window): min(len(signal), i + window + 1)]) for i in range(len(signal))])
    corrected = signal - baseline
    prominence = max(float(request.peak_prominence), float(np.nanmax(corrected)) * 0.02 if len(corrected) else 0.0)

    peak_spans: List[Dict[str, Any]] = []
    if analysis_engine == "MOCCA2":
        import mocca2
        mocca_peaks = mocca2.find_peaks(corrected, min_rel_height=max(float(request.peak_prominence), 0.0), min_height=0.0)
        for peak in mocca_peaks:
            idx = int(peak.maximum)
            peak_spans.append(
                {
                    "idx": idx,
                    "left": int(max(0, peak.left)),
                    "right": int(min(len(time) - 1, peak.right)),
                    "height": float(peak.height),
                    "prominence": float(getattr(peak, "prominence", peak.height)),
                    "peak_engine": "MOCCA2",
                }
            )
    else:
        peak_indices, props = find_peaks(corrected, prominence=prominence)
        widths = peak_widths(corrected, peak_indices, rel_height=0.5) if len(peak_indices) else ([], [], [], [])
        for rank, idx in enumerate(peak_indices):
            left = int(max(0, math.floor(widths[2][rank]))) if len(peak_indices) else int(idx)
            right = int(min(len(time) - 1, math.ceil(widths[3][rank]))) if len(peak_indices) else int(idx)
            peak_spans.append(
                {
                    "idx": int(idx),
                    "left": left,
                    "right": right,
                    "height": float(corrected[idx]),
                    "prominence": float(props.get("prominences", [None] * len(peak_indices))[rank]) if len(peak_indices) else None,
                    "peak_engine": "scipy.signal.find_peaks",
                }
            )

    total_positive_area = float(np.trapezoid(np.maximum(corrected, 0), time))
    peaks = []
    for rank, span in enumerate(peak_spans):
        idx = int(span["idx"])
        left = int(span["left"])
        right = int(span["right"])
        if right <= left:
            left = max(0, idx - 1)
            right = min(len(time) - 1, idx + 1)
        area = float(np.trapezoid(np.maximum(corrected[left : right + 1], 0), time[left : right + 1]))
        width_time = float(time[right] - time[left]) if right > left else 0.0
        height = float(span["height"])
        rt = float(time[idx])
        plates = 5.54 * (rt / width_time) ** 2 if width_time > 0 else None
        peaks.append(
            {
                "peak_id": rank + 1,
                "retention_time": rt,
                "area": area,
                "height": height,
                "width": width_time,
                "start_time": float(time[left]),
                "end_time": float(time[right]),
                "area_percent": (area / total_positive_area * 100.0) if total_positive_area else 0.0,
                "plates": plates,
                "resolution": None,
                "tailing_factor": None,
                "fit_model": request.fit_model,
                "peak_engine": span["peak_engine"],
                "prominence": span.get("prominence"),
            }
        )
    peaks.sort(key=lambda p: p["retention_time"])
    for i in range(1, len(peaks)):
        prev = peaks[i - 1]
        cur = peaks[i]
        denom = (prev.get("width") or 0) + (cur.get("width") or 0)
        cur["resolution"] = 2 * (cur["retention_time"] - prev["retention_time"]) / denom if denom else None
    return _json_clean(
        {
            "analysis_engine": analysis_engine,
            "engine_package": engine_package,
            "baseline_method": request.baseline_method,
            "fit_model": request.fit_model,
            "n_peaks": len(peaks),
            "peaks": peaks,
            "total_area": total_positive_area,
            "plotly_json": {
                "data": [
                    {"type": "scatter", "mode": "lines", "x": time.tolist(), "y": signal.tolist(), "name": "Raw"},
                    {"type": "scatter", "mode": "lines", "x": time.tolist(), "y": baseline.tolist(), "name": "Baseline"},
                    {"type": "scatter", "mode": "lines", "x": time.tolist(), "y": corrected.tolist(), "name": "Corrected"},
                ],
                "layout": {"title": "Chromatogram Analysis", "xaxis": {"title": "Time"}, "yaxis": {"title": "Signal"}},
            },
        }
    )


@router.post("/analysis/hplc/quick-analyze")
def chromatography_quick_analyze(request: ChromatographyAnalyzeRequest) -> Dict[str, Any]:
    return chromatography_analyze(request)


@router.post("/analysis/hplc/calibration-curve")
def hplc_calibration_curve(request: HplcCalibrationRequest) -> Dict[str, Any]:
    points = [p for p in request.points if math.isfinite(p.concentration) and math.isfinite(p.area)]
    if len(points) < 2:
        raise HTTPException(status_code=400, detail="At least two calibration points required")
    x = [p.concentration for p in points]
    y = [p.area for p in points]
    if request.force_through_origin:
        denom = sum(v * v for v in x)
        slope = sum(xi * yi for xi, yi in zip(x, y)) / denom if denom else 0.0
        intercept = 0.0
        fitted = [slope * xi for xi in x]
        ss_res = sum((yi - fi) ** 2 for yi, fi in zip(y, fitted))
        ss_tot = sum((yi - _mean(y)) ** 2 for yi in y)
        fit = {"slope": slope, "intercept": intercept, "r_squared": 1 - ss_res / ss_tot if ss_tot else 1.0, "p_value": None, "std_err": None}
    else:
        fit = _linear_fit(x, y)
    fit_x = sorted(x)
    fit_y = [fit["slope"] * value + fit["intercept"] for value in fit_x]
    return _json_clean(
        {
            **fit,
            "analyte_name": request.analyte_name,
            "unit": request.unit,
            "n_points": len(points),
            "points": [p.model_dump() for p in points],
            "plotly_json": {
                "data": [
                    {
                        "type": "scatter",
                        "mode": "markers",
                        "x": x,
                        "y": y,
                        "name": "Calibration standards",
                    },
                    {
                        "type": "scatter",
                        "mode": "lines",
                        "x": fit_x,
                        "y": fit_y,
                        "name": "Linear fit",
                    },
                ],
                "layout": {
                    "title": f"HPLC Calibration - {request.analyte_name or 'Analyte'}",
                    "xaxis": {"title": f"Concentration ({request.unit})"},
                    "yaxis": {"title": "Peak area"},
                },
            },
        }
    )


@router.post("/analysis/hplc/quantify")
def hplc_quantify(request: HplcQuantifyRequest) -> Dict[str, Any]:
    if request.cal_concentrations and request.cal_areas:
        if len(request.cal_concentrations) != len(request.cal_areas):
            raise HTTPException(status_code=400, detail="cal_concentrations and cal_areas must have the same length")
        curve = hplc_calibration_curve(
            HplcCalibrationRequest(
                points=[CalibrationPoint(concentration=c, area=a) for c, a in zip(request.cal_concentrations, request.cal_areas)],
                unit=request.unit,
            )
        )
        slope = float(curve["slope"])
        intercept = float(curve["intercept"])
        if slope == 0:
            raise HTTPException(status_code=400, detail="Calibration slope cannot be zero")
        sample_areas = []
        for area in request.sample_areas:
            parsed_area = _finite_float(area)
            if parsed_area is None:
                raise HTTPException(status_code=400, detail="sample_areas must contain only finite values")
            sample_areas.append(parsed_area)
        sample_ids = _required_ids(request.sample_ids, len(sample_areas), "sample_ids")
        sample_rows = []
        for sample_id, a in zip(sample_ids, sample_areas):
            concentration = (a - intercept) / slope * request.dilution_factor
            sample_rows.append({"id": sample_id, "area": a, "concentration": concentration, "unit": request.unit})
        return _json_clean(
            {
                "curve_stats": curve,
                "samples": sample_rows,
                "summary": f"Quantified {len(sample_rows)} HPLC sample(s) against {len(request.cal_concentrations)} calibration levels.",
                "plotly_json": {
                    "data": [
                        {"type": "scatter", "mode": "markers", "x": request.cal_concentrations, "y": request.cal_areas, "name": "Calibration"},
                        {
                            "type": "scatter",
                            "mode": "lines",
                            "x": request.cal_concentrations,
                            "y": [slope * x + intercept for x in request.cal_concentrations],
                            "name": "Fit",
                        },
                    ],
                    "layout": {"title": "HPLC Calibration", "xaxis": {"title": f"Concentration ({request.unit})"}, "yaxis": {"title": "Area"}},
                },
            }
        )

    if request.area is None or request.slope is None:
        raise HTTPException(status_code=400, detail="area and slope are required unless calibration/sample arrays are provided")
    if request.slope == 0:
        raise HTTPException(status_code=400, detail="Slope cannot be zero")
    concentration = (request.area - request.intercept) / request.slope
    return _json_clean({"area": request.area, "concentration": concentration, "dilution_corrected_concentration": concentration * request.dilution_factor})


@router.post("/analysis/hplc/plasmid/isoforms")
@router.post("/analysis/hplc/empower/plasmid-isoforms")
def chromatography_plasmid_isoforms(request: PlasmidIsoformRequest) -> Dict[str, Any]:
    if not request.peaks:
        raise HTTPException(status_code=400, detail="peaks must contain real peak rows with area and retention_time fields")
    if not request.windows:
        raise HTTPException(status_code=400, detail="windows must explicitly define retention-time bounds for each plasmid isoform")

    peak_rows = []
    for idx, peak in enumerate(request.peaks, start=1):
        area = _require_finite(peak, ("area",), "Area", idx, "plasmid isoform peak")
        rt = _require_finite(peak, ("retention_time", "retention time", "rt"), "Retention Time", idx, "plasmid isoform peak")
        peak_rows.append({"area": area, "retention_time": rt})

    window_rows: Dict[str, Tuple[float, float]] = {}
    for name, bounds in request.windows.items():
        isoform_name = str(name).strip()
        if not isoform_name:
            raise HTTPException(status_code=400, detail="Each isoform window must have a non-empty name")
        lo, hi = float(bounds[0]), float(bounds[1])
        if not math.isfinite(lo) or not math.isfinite(hi) or hi < lo:
            raise HTTPException(status_code=400, detail=f"Isoform window {isoform_name} must provide finite [low, high] bounds")
        window_rows[isoform_name] = (lo, hi)

    total_area = sum(p["area"] for p in peak_rows)
    if total_area <= 0:
        raise HTTPException(status_code=400, detail="Peak areas must sum to a positive value for plasmid isoform analysis")
    isoforms: Dict[str, Dict[str, Any]] = {}
    assigned = 0.0
    for name, (lo, hi) in window_rows.items():
        matched = [p for p in peak_rows if lo <= p["retention_time"] <= hi]
        area = sum(p["area"] for p in matched)
        assigned += area
        isoforms[name] = {"area": area, "area_percent": area / total_area * 100.0, "peak_count": len(matched), "window": [lo, hi]}
    return _json_clean({"isoforms": isoforms, "total_area": total_area, "assigned_area": assigned, "total_assigned_area_percent": assigned / total_area * 100.0})


# Minimal in-memory Empower review cache for UI sessions. This keeps BMS source-truth local;
# regulated Empower remains the acquisition/integration source of truth.
_EMPOWER_IMPORTS: Dict[int, Dict[str, Any]] = {}
_NEXT_IMPORT_ID = 1
_EMPOWER_TEXT_EXPORT_EXTENSIONS = {'csv', 'txt'}
_EMPOWER_PARSEABLE_NATIVE_EXTENSIONS = {'arw', 'cdf', 'zip'}
_EMPOWER_NATIVE_DATABASE_EXTENSIONS = {'raw', 'dat', 'mdb', 'accdb', 'db'}


def _upload_extension(filename: Optional[str]) -> str:
    if not filename or '.' not in filename:
        return ''
    return filename.rsplit('.', 1)[-1].strip().lower()


def _sst_summary(injections: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    by_group: Dict[str, List[Dict[str, Any]]] = {}
    for inj in injections:
        if inj.get("is_excluded"):
            continue
        sample_type = inj.get("sample_type")
        if sample_type is None:
            continue
        by_group.setdefault(str(sample_type), []).append(inj)
    out = []
    for group, rows in sorted(by_group.items()):
        areas = [float(r["primary_peak_area"] if r.get("primary_peak_area") is not None else r["total_area"]) for r in rows if _finite_float(r.get("primary_peak_area") if r.get("primary_peak_area") is not None else r.get("total_area")) is not None]
        pcts = [float(r.get("primary_peak_percent")) for r in rows if _finite_float(r.get("primary_peak_percent")) is not None]
        rts = [float(r.get("primary_peak_rt")) for r in rows if _finite_float(r.get("primary_peak_rt")) is not None]
        res = [float(r.get("primary_peak_resolution")) for r in rows if _finite_float(r.get("primary_peak_resolution")) is not None]
        out.append(
            {
                "sst_group": group,
                "n_injections": len(rows),
                "area_mean": _mean(areas),
                "area_rsd": _cv_percent(areas),
                "percent_primary_mean": _mean(pcts),
                "percent_primary_rsd": _cv_percent(pcts),
                "rt_mean": _mean(rts),
                "rt_rsd": _cv_percent(rts),
                "resolution_mean": _mean(res),
                "resolution_rsd": _cv_percent(res),
            }
        )
    return _json_clean(out)


def _infer_empower_sample_role(sample_name: Any, sample_type: Any) -> Tuple[str, str]:
    """Return a conservative role label derived only from real export metadata."""
    sample_type_text = str(sample_type or "").strip().upper()
    if sample_type_text in {"SST", "STANDARD", "SAMPLE", "BLANK", "CONTROL"}:
        return sample_type_text.lower(), "sample_type"
    name = str(sample_name or "").strip().lower()
    if re.search(r"(^|[_\s-])sst([_\s-]|$)", name):
        return "sst", "sample_name_pattern"
    if "blank" in name:
        return "blank", "sample_name_pattern"
    if re.search(r"(^|[_\s-])(std|standard)([_\s-]|$)", name):
        return "standard", "sample_name_pattern"
    if re.search(r"(^|[_\s-])(ctrl|control)([_\s-]|$)", name):
        return "control", "sample_name_pattern"
    return "sample", "default_untyped_export"


def _safe_median(values: List[float]) -> Optional[float]:
    return float(statistics.median(values)) if values else None


def _empower_peak_table(injections: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for injection in injections:
        primary_area = _finite_float(injection.get("primary_peak_area"))
        for peak in injection.get("peaks", []) or []:
            area = _finite_float(peak.get("area"))
            rows.append(
                {
                    "import_id": injection.get("import_id"),
                    "injection_id": injection.get("id"),
                    "injection_number": injection.get("injection_number"),
                    "sample_name": injection.get("sample_name"),
                    "sample_type": injection.get("sample_type"),
                    "sample_role": injection.get("sample_role"),
                    "source_file": injection.get("source_file"),
                    "peak_id": peak.get("peak_id"),
                    "retention_time": peak.get("retention_time"),
                    "retention_time_min": peak.get("retention_time_min", peak.get("retention_time")),
                    "area": area,
                    "height": peak.get("height"),
                    "area_percent": peak.get("area_percent"),
                    "resolution": peak.get("resolution"),
                    "tailing_factor": peak.get("tailing_factor"),
                    "peak_source": peak.get("peak_source"),
                    "is_primary_peak": bool(primary_area is not None and area is not None and math.isclose(area, primary_area, rel_tol=1e-9, abs_tol=1e-9)),
                }
            )
    return _json_clean(rows)


def _empower_peak_region_summary(injections: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    summaries: List[Dict[str, Any]] = []
    for injection in injections:
        peaks = [peak for peak in (injection.get("peaks", []) or []) if _finite_float(peak.get("area")) is not None and _finite_float(peak.get("retention_time")) is not None]
        total_area = sum(float(peak["area"]) for peak in peaks)
        primary = max(peaks, key=lambda peak: float(peak["area"])) if peaks else None
        primary_rt = _finite_float(primary.get("retention_time")) if primary else None
        primary_area = _finite_float(primary.get("area")) if primary else None
        pre_area = 0.0
        post_area = 0.0
        pre_count = 0
        post_count = 0
        if primary is not None and primary_rt is not None:
            primary_peak_id = primary.get("peak_id")
            for peak in peaks:
                if peak.get("peak_id") == primary_peak_id:
                    continue
                rt = float(peak["retention_time"])
                area = float(peak["area"])
                if rt < primary_rt:
                    pre_area += area
                    pre_count += 1
                elif rt > primary_rt:
                    post_area += area
                    post_count += 1
        summaries.append(
            {
                "import_id": injection.get("import_id"),
                "injection_id": injection.get("id"),
                "injection_number": injection.get("injection_number"),
                "sample_name": injection.get("sample_name"),
                "sample_role": injection.get("sample_role"),
                "primary_peak_rt": primary_rt,
                "total_area": total_area,
                "pre_primary_area": pre_area,
                "primary_area": primary_area,
                "post_primary_area": post_area,
                "pre_primary_area_percent": pre_area / total_area * 100.0 if total_area else 0.0,
                "primary_area_percent": primary_area / total_area * 100.0 if total_area and primary_area is not None else None,
                "post_primary_area_percent": post_area / total_area * 100.0 if total_area else 0.0,
                "pre_primary_peak_count": pre_count,
                "post_primary_peak_count": post_count,
                "peak_count": len(peaks),
            }
        )
    return _json_clean(summaries)


def _empower_batch_analytics(injections: List[Dict[str, Any]], source_format_counts: Dict[str, int]) -> Dict[str, Any]:
    peak_table = _empower_peak_table(injections)
    peak_region_summary = _empower_peak_region_summary(injections)
    primary_rts = [float(inj["primary_peak_rt"]) for inj in injections if _finite_float(inj.get("primary_peak_rt")) is not None]
    primary_pcts = [float(inj["primary_peak_percent"]) for inj in injections if _finite_float(inj.get("primary_peak_percent")) is not None]
    total_areas = [float(inj["total_area"]) for inj in injections if _finite_float(inj.get("total_area")) is not None]
    median_rt = _safe_median(primary_rts)
    median_area = _safe_median(total_areas)
    flag_counts: Dict[str, int] = {}
    flagged_injection_count = 0
    for injection in injections:
        flags: List[str] = []
        pct = _finite_float(injection.get("primary_peak_percent"))
        rt = _finite_float(injection.get("primary_peak_rt"))
        area = _finite_float(injection.get("total_area"))
        peak_count = len(injection.get("peaks", []) or [])
        if pct is not None and pct < 90.0:
            flags.append("low_primary_percent")
        if median_rt is not None and rt is not None and abs(rt - median_rt) > 0.35:
            flags.append("rt_drift")
        if median_area is not None and area is not None and area < median_area * 0.20:
            flags.append("low_total_area")
        if peak_count >= 4:
            flags.append("complex_peak_pattern")
        injection["qc_flags"] = flags
        injection["peak_count"] = peak_count
        if flags:
            flagged_injection_count += 1
            for flag in flags:
                flag_counts[flag] = flag_counts.get(flag, 0) + 1
    sample_role_counts: Dict[str, int] = {}
    method_counts: Dict[str, int] = {}
    injection_volume_counts: Dict[str, int] = {}
    dates = [str(inj.get("run_date")) for inj in injections if inj.get("run_date")]
    for injection in injections:
        role = str(injection.get("sample_role") or "unknown")
        sample_role_counts[role] = sample_role_counts.get(role, 0) + 1
        method = str(injection.get("method_name") or "UNSPECIFIED_BY_EXPORT")
        method_counts[method] = method_counts.get(method, 0) + 1
        volume = _finite_float(injection.get("injection_volume"))
        volume_key = f"{volume:g}" if volume is not None else "UNSPECIFIED_BY_EXPORT"
        injection_volume_counts[volume_key] = injection_volume_counts.get(volume_key, 0) + 1
    summary = {
        "n_injections": len(injections),
        "n_chromatograms": sum(1 for inj in injections if isinstance(inj.get("chromatogram"), dict)),
        "total_peak_rows": len(peak_table),
        "native_peak_rows": sum(int(inj.get("native_peak_count") or 0) for inj in injections),
        "source_format_counts": source_format_counts,
        "sample_role_counts": sample_role_counts,
        "method_counts": method_counts,
        "injection_volume_counts": injection_volume_counts,
        "run_date_min": min(dates) if dates else None,
        "run_date_max": max(dates) if dates else None,
        "primary_rt_median": median_rt,
        "primary_rt_mean": _mean(primary_rts),
        "primary_rt_rsd": _cv_percent(primary_rts),
        "primary_percent_mean": _mean(primary_pcts),
        "primary_percent_rsd": _cv_percent(primary_pcts),
        "total_area_mean": _mean(total_areas),
        "total_area_rsd": _cv_percent(total_areas),
        "flagged_injection_count": flagged_injection_count,
        "flag_counts": flag_counts,
        "role_source_note": "sample_role is derived from Empower sample_type when present, otherwise from real sample_name tokens such as SST or BLANK; it is not an invented assay row.",
    }
    return {"empower_summary": _json_clean(summary), "peak_table": peak_table, "peak_region_summary": peak_region_summary}


def _empower_qc_plot(injections: List[Dict[str, Any]]) -> Dict[str, Any]:
    x = [inj.get("id") for inj in injections]
    labels = [f"{inj.get('id')}: {inj.get('sample_name')}" for inj in injections]
    return _json_clean(
        {
            "data": [
                {"type": "scatter", "mode": "lines+markers", "x": x, "y": [inj.get("primary_peak_percent") for inj in injections], "text": labels, "name": "Primary % area", "line": {"color": "#38bdf8", "width": 2}, "marker": {"size": 7}, "hovertemplate": "%{text}<br>Primary area %{y:.2f}%<extra></extra>"},
                {"type": "scatter", "mode": "lines+markers", "x": x, "y": [inj.get("primary_peak_rt") for inj in injections], "text": labels, "name": "Primary RT", "line": {"color": "#a78bfa", "width": 2}, "marker": {"size": 7}, "yaxis": "y2", "hovertemplate": "%{text}<br>Primary RT %{y:.3f} min<extra></extra>"},
                {"type": "bar", "x": x, "y": [inj.get("total_area") for inj in injections], "text": labels, "name": "Total area", "marker": {"color": "rgba(52, 211, 153, 0.45)"}, "yaxis": "y3", "hovertemplate": "%{text}<br>Total area %{y:.3s}<extra></extra>"},
                {"type": "scatter", "mode": "markers", "x": x, "y": [inj.get("peak_count") for inj in injections], "text": labels, "name": "Peak count", "marker": {"color": "#fbbf24", "size": 9, "symbol": "square"}, "yaxis": "y4", "hovertemplate": "%{text}<br>Peak count %{y}<extra></extra>"},
            ],
            "layout": {
                "title": "Empower Batch QC",
                "xaxis": {"title": "Injection row", "domain": [0.0, 0.92]},
                "yaxis": {"title": "Primary area %", "range": [0, 105]},
                "yaxis2": {"title": "Primary RT (min)", "overlaying": "y", "side": "right", "showgrid": False},
                "yaxis3": {"title": "Total area", "overlaying": "y", "side": "right", "position": 0.96, "showgrid": False, "visible": False},
                "yaxis4": {"title": "Peak count", "overlaying": "y", "side": "left", "showgrid": False, "visible": False},
                "hovermode": "closest",
                "legend": {"orientation": "h", "y": -0.25},
                "margin": {"l": 60, "r": 90, "t": 70, "b": 90},
            },
        }
    )


def _empower_composition_plot(peak_region_summary: List[Dict[str, Any]]) -> Dict[str, Any]:
    x = [row.get("injection_id") for row in peak_region_summary]
    labels = [f"{row.get('injection_id')}: {row.get('sample_name')}" for row in peak_region_summary]
    return _json_clean(
        {
            "data": [
                {"type": "bar", "x": x, "y": [row.get("pre_primary_area_percent") for row in peak_region_summary], "text": labels, "name": "Pre-primary area %", "marker": {"color": "#f472b6"}, "hovertemplate": "%{text}<br>Pre-primary %{y:.2f}%<extra></extra>"},
                {"type": "bar", "x": x, "y": [row.get("primary_area_percent") for row in peak_region_summary], "text": labels, "name": "Primary area %", "marker": {"color": "#38bdf8"}, "hovertemplate": "%{text}<br>Primary %{y:.2f}%<extra></extra>"},
                {"type": "bar", "x": x, "y": [row.get("post_primary_area_percent") for row in peak_region_summary], "text": labels, "name": "Post-primary area %", "marker": {"color": "#fb923c"}, "hovertemplate": "%{text}<br>Post-primary %{y:.2f}%<extra></extra>"},
            ],
            "layout": {"title": "Empower Peak Composition by Injection", "barmode": "stack", "xaxis": {"title": "Injection row"}, "yaxis": {"title": "Area % of native peak table", "range": [0, 100]}, "legend": {"orientation": "h", "y": -0.25}, "hovermode": "closest", "margin": {"l": 60, "r": 30, "t": 70, "b": 90}},
        }
    )


def _decode_netcdf_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, bytes):
        text = value.decode("utf-8", errors="ignore")
    elif isinstance(value, np.ndarray):
        if value.dtype.kind in {"S", "U"}:
            text = b"".join(value.astype("S").flat).decode("utf-8", errors="ignore")
        else:
            text = str(value.tolist())
    else:
        text = str(value)
    text = text.replace("\x00", "").strip()
    return text if text else None


def _netcdf_attr(dataset: Any, name: str) -> Optional[str]:
    return _decode_netcdf_text(getattr(dataset, name, None))


def _netcdf_array(dataset: Any, name: str) -> np.ndarray:
    variable = dataset.variables.get(name)
    if variable is None:
        return np.asarray([], dtype=float)
    return np.asarray(variable.data).copy()


def _valid_chrom_float(value: Any) -> Optional[float]:
    parsed = _finite_float(value)
    if parsed is None or abs(parsed) >= 1e30:
        return None
    return parsed


def _downsample_xy(x: np.ndarray, y: np.ndarray, max_points: int = 1200) -> Dict[str, Any]:
    if len(x) != len(y):
        n = min(len(x), len(y))
        x = x[:n]
        y = y[:n]
    if len(x) <= max_points:
        idx = np.arange(len(x))
    else:
        idx = np.unique(np.linspace(0, len(x) - 1, max_points).astype(int))
    return {"time_min": x[idx].astype(float).tolist(), "signal": y[idx].astype(float).tolist(), "points": int(len(x)), "downsampled": bool(len(x) > len(idx))}


def _empower_base_name(filename: str) -> str:
    base = filename.replace("\\", "/").rsplit("/", 1)[-1]
    return base.rsplit(".", 1)[0]


def _empower_parse_comment(comment: Optional[str]) -> Dict[str, Optional[str]]:
    if not comment:
        return {"vial": None, "injection_number": None}
    vial_match = re.search(r"vial\s+([^\s]+)", comment, flags=re.IGNORECASE)
    injection_match = re.search(r"injection\s+(\d+)", comment, flags=re.IGNORECASE)
    return {
        "vial": vial_match.group(1).strip() if vial_match else None,
        "injection_number": injection_match.group(1).strip() if injection_match else None,
    }


def _empower_native_peak_rows(retention_unit: Optional[str], rt_values: np.ndarray, area_values: np.ndarray, height_values: np.ndarray) -> List[Dict[str, Any]]:
    peaks: List[Dict[str, Any]] = []
    count = max(len(rt_values), len(area_values), len(height_values))
    for idx in range(count):
        rt = _valid_chrom_float(rt_values[idx]) if idx < len(rt_values) else None
        area = _valid_chrom_float(area_values[idx]) if idx < len(area_values) else None
        height = _valid_chrom_float(height_values[idx]) if idx < len(height_values) else None
        if rt is None or area is None or area <= 0:
            continue
        if (retention_unit or "").lower().startswith("sec"):
            rt_sec = rt
            rt_min = rt / 60.0
        else:
            rt_min = rt
            rt_sec = rt * 60.0
        peaks.append(
            {
                "peak_id": len(peaks) + 1,
                "retention_time": rt_min,
                "retention_time_min": rt_min,
                "retention_time_sec": rt_sec,
                "area": area,
                "height": height,
                "area_percent": None,
                "peak_source": "empower_aia_cdf_native_peak_table",
                "resolution": None,
                "tailing_factor": None,
            }
        )
    total = sum(float(peak["area"]) for peak in peaks)
    if total > 0:
        for peak in peaks:
            peak["area_percent"] = float(peak["area"]) / total * 100.0
    return peaks


def _empower_injection_from_cdf(payload: bytes, filename: str) -> Dict[str, Any]:
    try:
        from scipy.io import netcdf_file
    except ImportError as exc:  # pragma: no cover - scipy is required by this API image
        raise HTTPException(status_code=500, detail="scipy.io.netcdf_file is not available for Empower AIA CDF parsing") from exc
    try:
        with netcdf_file(io.BytesIO(payload), "r", mmap=False) as dataset:
            sample_name = _netcdf_attr(dataset, "sample_name")
            sample_type = (_netcdf_attr(dataset, "sample_type") or "UNSPECIFIED_BY_EXPORT").upper()
            injection_date = _netcdf_attr(dataset, "injection_date_time_stamp")
            injection_volume = _valid_chrom_float(_netcdf_attr(dataset, "sample_injection_volume"))
            detector_unit = _netcdf_attr(dataset, "detector_unit")
            retention_unit = _netcdf_attr(dataset, "retention_unit") or "seconds"
            detector_name = _netcdf_attr(dataset, "detector_name")
            comment = _netcdf_attr(dataset, "sample_id_comments")
            parsed_comment = _empower_parse_comment(comment)
            raw_time = _netcdf_array(dataset, "raw_data_retention").astype(float)
            signal = _netcdf_array(dataset, "ordinate_values").astype(float)
            if signal.size == 0:
                raise HTTPException(status_code=400, detail=f"{filename}: AIA CDF did not contain ordinate_values signal array")
            if raw_time.size == 0:
                sampling_interval = _valid_chrom_float(_netcdf_array(dataset, "actual_sampling_interval").reshape(-1)[0] if _netcdf_array(dataset, "actual_sampling_interval").size else None)
                delay_time = _valid_chrom_float(_netcdf_array(dataset, "actual_delay_time").reshape(-1)[0] if _netcdf_array(dataset, "actual_delay_time").size else 0.0)
                run_time = _valid_chrom_float(_netcdf_array(dataset, "actual_run_time_length").reshape(-1)[0] if _netcdf_array(dataset, "actual_run_time_length").size else None)
                if sampling_interval is None or sampling_interval <= 0:
                    if run_time is None or run_time <= 0 or signal.size < 2:
                        raise HTTPException(status_code=400, detail=f"{filename}: AIA CDF omitted raw_data_retention and lacks usable sampling interval/run time metadata")
                    sampling_interval = run_time / float(max(signal.size - 1, 1))
                raw_time = (delay_time or 0.0) + np.arange(signal.size, dtype=float) * sampling_interval
            n = min(raw_time.size, signal.size)
            raw_time = raw_time[:n]
            signal = signal[:n]
            valid = np.isfinite(raw_time) & np.isfinite(signal) & (np.abs(raw_time) < 1e30) & (np.abs(signal) < 1e30)
            raw_time = raw_time[valid]
            signal = signal[valid]
            if raw_time.size < 3:
                raise HTTPException(status_code=400, detail=f"{filename}: AIA CDF chromatogram has fewer than 3 finite points")
            time_min = raw_time / 60.0 if retention_unit.lower().startswith("sec") else raw_time
            peaks = _empower_native_peak_rows(
                retention_unit,
                _netcdf_array(dataset, "peak_retention_time").astype(float),
                _netcdf_array(dataset, "peak_area").astype(float),
                _netcdf_array(dataset, "peak_height").astype(float),
            )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"{filename}: unable to parse Empower AIA CDF chromatogram: {exc}") from exc

    total_area = sum(float(peak["area"]) for peak in peaks)
    primary = max(peaks, key=lambda peak: float(peak["area"])) if peaks else None
    return {
        "sample_name": sample_name or _empower_base_name(filename),
        "sample_type": sample_type,
        "injection_number": parsed_comment.get("injection_number"),
        "vial": parsed_comment.get("vial"),
        "method_name": None,
        "run_date": injection_date,
        "source_file": filename,
        "source_format": "empower_aia_cdf",
        "chromatogram_points": int(time_min.size),
        "chromatogram": _downsample_xy(time_min, signal),
        "detector_name": detector_name,
        "detector_unit": detector_unit,
        "retention_unit": "minutes",
        "injection_volume": injection_volume,
        "peaks": peaks,
        "native_peak_count": len(peaks),
        "total_area": total_area,
        "primary_peak_area": primary.get("area") if primary else None,
        "primary_peak_percent": primary.get("area_percent") if primary else None,
        "primary_peak_rt": primary.get("retention_time") if primary else None,
        "primary_peak_resolution": primary.get("resolution") if primary else None,
        "is_excluded": False,
        "note": None,
        "flag": None,
        "import_engine": "scipy.io.netcdf_file",
    }


def _empower_arw_metadata_and_chromatogram(payload: bytes, filename: str) -> Dict[str, Any]:
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = payload.decode("latin-1", errors="ignore")
    lines = text.splitlines()
    if len(lines) < 3:
        raise HTTPException(status_code=400, detail=f"{filename}: ARW export has no chromatogram rows")
    header = next(csv.reader([lines[0]], delimiter="\t"))
    values = next(csv.reader([lines[1]], delimiter="\t"))
    meta = {header[idx].strip(): values[idx].strip() if idx < len(values) else "" for idx in range(len(header)) if header[idx].strip()}
    time_values: List[float] = []
    signal_values: List[float] = []
    for line in lines[2:]:
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        time_value = _finite_float(parts[0])
        signal_value = _finite_float(parts[1])
        if time_value is None or signal_value is None:
            continue
        time_values.append(time_value)
        signal_values.append(signal_value)
    if len(time_values) < 3:
        raise HTTPException(status_code=400, detail=f"{filename}: ARW export has fewer than 3 finite chromatogram points")
    return {"metadata": meta, "time_min": np.asarray(time_values, dtype=float), "signal": np.asarray(signal_values, dtype=float), "filename": filename}


def _empower_injection_from_arw(payload: bytes, filename: str, baseline_method: str, peak_prominence: float) -> Dict[str, Any]:
    parsed = _empower_arw_metadata_and_chromatogram(payload, filename)
    meta = parsed["metadata"]
    analysis = chromatography_analyze(
        ChromatographyAnalyzeRequest(
            time=parsed["time_min"].astype(float).tolist(),
            signal=parsed["signal"].astype(float).tolist(),
            baseline_method=baseline_method,
            peak_prominence=peak_prominence,
        )
    )
    peaks = analysis.get("peaks", [])
    total_area = float(analysis.get("total_area") or 0.0)
    primary = max(peaks, key=lambda peak: float(peak["area"])) if peaks else None
    return {
        "sample_name": meta.get("SampleName") or _empower_base_name(filename),
        "sample_type": (meta.get("Sample Type") or "UNSPECIFIED_BY_EXPORT").upper(),
        "injection_number": meta.get("Injection") or None,
        "vial": meta.get("Vial") or None,
        "method_name": meta.get("Acq Method Set") or None,
        "run_date": meta.get("Date Acquired") or None,
        "source_file": filename,
        "source_format": "empower_arw_text_chromatogram",
        "chromatogram_points": int(len(parsed["time_min"])),
        "chromatogram": _downsample_xy(parsed["time_min"], parsed["signal"]),
        "detector_name": meta.get("Channel") or None,
        "detector_unit": "AU",
        "retention_unit": "minutes",
        "injection_volume": _valid_chrom_float(meta.get("Injection Volume")),
        "peaks": peaks,
        "native_peak_count": 0,
        "total_area": total_area,
        "primary_peak_area": primary.get("area") if primary else None,
        "primary_peak_percent": primary.get("area_percent") if primary else None,
        "primary_peak_rt": primary.get("retention_time") if primary else None,
        "primary_peak_resolution": primary.get("resolution") if primary else None,
        "is_excluded": False,
        "note": None,
        "flag": None,
        "import_engine": analysis.get("analysis_engine") or "scipy.signal",
    }


def _merge_arw_metadata(record: Dict[str, Any], arw: Dict[str, Any]) -> Dict[str, Any]:
    meta = arw.get("metadata", {})
    record["paired_arw_file"] = arw.get("filename")
    record["method_name"] = record.get("method_name") or meta.get("Acq Method Set") or None
    record["system_name"] = meta.get("System Name") or None
    record["acquired_by"] = meta.get("Acquired By") or None
    record["instrument_method_id"] = meta.get("Instrument Method Id") or None
    record["channel_description"] = meta.get("Channel Description") or None
    if not record.get("injection_number"):
        record["injection_number"] = meta.get("Injection") or None
    if not record.get("vial"):
        record["vial"] = meta.get("Vial") or None
    if not record.get("run_date"):
        record["run_date"] = meta.get("Date Acquired") or None
    if not record.get("injection_volume"):
        record["injection_volume"] = _valid_chrom_float(meta.get("Injection Volume"))
    return record



def _empower_records_from_grouped_cdf_arw(
    grouped: Dict[str, Dict[str, Tuple[str, bytes]]],
    baseline_method: str,
    peak_prominence: float,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Parse CDF/ARW groups from ZIP imports or browser folder uploads.

    Empower mass-download exports commonly contain one AIA CDF and one ARW text
    chromatogram per injection. Browser folder uploads send those as many
    independent multipart files; the importer must pair them by basename instead
    of emitting duplicate CDF and ARW injections.
    """
    records: List[Dict[str, Any]] = []
    errors: List[str] = []
    for _base, members in sorted(grouped.items()):
        arw_parsed: Optional[Dict[str, Any]] = None
        if "arw" in members:
            arw_name, arw_payload = members["arw"]
            try:
                arw_parsed = _empower_arw_metadata_and_chromatogram(arw_payload, arw_name)
            except HTTPException as exc:
                errors.append(str(exc.detail))
        if "cdf" in members:
            cdf_name, cdf_payload = members["cdf"]
            try:
                record = _empower_injection_from_cdf(cdf_payload, cdf_name)
                if arw_parsed is not None:
                    record = _merge_arw_metadata(record, arw_parsed)
                records.append(record)
            except HTTPException as exc:
                errors.append(str(exc.detail))
        elif "arw" in members:
            arw_name, arw_payload = members["arw"]
            try:
                records.append(_empower_injection_from_arw(arw_payload, arw_name, baseline_method, peak_prominence))
            except HTTPException as exc:
                errors.append(str(exc.detail))
    return records, errors

def _empower_records_from_zip(payload: bytes, filename: str, baseline_method: str, peak_prominence: float) -> Tuple[List[Dict[str, Any]], List[str]]:
    try:
        archive = zipfile.ZipFile(io.BytesIO(payload))
    except zipfile.BadZipFile as exc:
        raise HTTPException(status_code=400, detail=f"{filename}: not a valid ZIP archive: {exc}") from exc
    records: List[Dict[str, Any]] = []
    errors: List[str] = []
    with archive:
        grouped: Dict[str, Dict[str, Tuple[str, bytes]]] = {}
        for member in sorted(archive.namelist()):
            if member.endswith("/"):
                continue
            ext = _upload_extension(member)
            if ext not in {"cdf", "arw"}:
                continue
            grouped.setdefault(_empower_base_name(member), {})[ext] = (member, archive.read(member))
        if not grouped:
            raise HTTPException(status_code=400, detail=f"{filename}: ZIP did not contain parseable Empower .cdf or .arw chromatogram files")
        records, errors = _empower_records_from_grouped_cdf_arw(grouped, baseline_method, peak_prominence)
    return records, errors


def _finalize_empower_injection(record: Dict[str, Any], import_id: int, injections: List[Dict[str, Any]]) -> Dict[str, Any]:
    finalized = {**record}
    finalized["id"] = len(injections) + 1
    finalized["import_id"] = import_id
    finalized.setdefault("is_excluded", False)
    finalized.setdefault("note", None)
    finalized.setdefault("flag", None)
    finalized.setdefault("sample_type", "UNKNOWN")
    finalized.setdefault("injection_number", str(finalized["id"]))
    sample_role, sample_role_source = _infer_empower_sample_role(finalized.get("sample_name"), finalized.get("sample_type"))
    finalized["sample_role"] = sample_role
    finalized["sample_role_source"] = sample_role_source
    finalized["peak_count"] = len(finalized.get("peaks", []) or [])
    finalized.setdefault("qc_flags", [])
    return _json_clean(finalized)


def _empower_chromatogram_plot(injections: List[Dict[str, Any]]) -> Dict[str, Any]:
    traces = []
    for injection in injections:
        chromatogram = injection.get("chromatogram")
        if not isinstance(chromatogram, dict):
            continue
        traces.append(
            {
                "type": "scatter",
                "mode": "lines",
                "x": chromatogram.get("time_min", []),
                "y": chromatogram.get("signal", []),
                "name": f"{injection.get('id')}: {injection.get('sample_name')}",
                "opacity": 0.55,
            }
        )
    return {"data": traces, "layout": {"title": "Empower Chromatograms", "xaxis": {"title": "Time (min)"}, "yaxis": {"title": "Signal (AU)"}}}


@router.post("/analysis/hplc/empower/import")
async def empower_import(
    files: List[UploadFile] = File(...),
    persist: bool = Form(True),
    baseline_method: str = Form("rolling_min"),
    peak_prominence: float = Form(0.05),
) -> Dict[str, Any]:
    global _NEXT_IMPORT_ID
    import_id = _NEXT_IMPORT_ID
    _NEXT_IMPORT_ID += 1
    injections: List[Dict[str, Any]] = []
    errors: List[str] = []
    source_files: List[Dict[str, Any]] = []
    loose_cdf_arw: Dict[str, Dict[str, Tuple[str, bytes]]] = {}
    for file_idx, file in enumerate(files, start=1):
        extension = _upload_extension(file.filename)
        payload = await file.read()
        source_files.append(
            {
                "index": file_idx,
                "filename": file.filename or f"upload_{file_idx}.{extension or 'dat'}",
                "content_type": file.content_type,
                "size_bytes": len(payload),
                "sha256": hashlib.sha256(payload).hexdigest(),
                "content_bytes": payload,
                "original_relative_path": file.filename or f"upload_{file_idx}.{extension or 'dat'}",
            }
        )
        source_file_record = source_files[-1]
        if extension == "zip":
            try:
                with zipfile.ZipFile(io.BytesIO(payload)) as archive:
                    archive_members = []
                    for member in sorted(archive.namelist()):
                        if member.endswith("/"):
                            continue
                        member_ext = _upload_extension(member)
                        if member_ext not in {"cdf", "arw"}:
                            continue
                        archive_members.append(
                            {
                                "member_name": member,
                                "extension": member_ext,
                                "member_role": "empower_aia_cdf" if member_ext == "cdf" else "empower_arw_metadata",
                                "content_bytes": archive.read(member),
                            }
                        )
                    if archive_members:
                        source_file_record["archive_members"] = archive_members
            except zipfile.BadZipFile:
                pass
        if extension in _EMPOWER_NATIVE_DATABASE_EXTENSIONS:
            errors.append(
                f"{file.filename}: Native Empower database/RAW files are not currently parsed by BMS for extension .{extension}; upload Empower AIA .cdf, ARW chromatogram text, ZIP containing .cdf/.arw, or CSV/ASCII injection/peak export"
            )
            continue
        if extension == "zip":
            try:
                records, native_errors = _empower_records_from_zip(payload, file.filename or f"upload_{file_idx}.zip", baseline_method, peak_prominence)
            except HTTPException as exc:
                errors.append(str(exc.detail))
                continue
            errors.extend(native_errors)
            for record in records:
                injections.append(_finalize_empower_injection(record, import_id, injections))
            continue
        if extension in {"cdf", "arw"}:
            upload_name = file.filename or f"upload_{file_idx}.{extension}"
            loose_cdf_arw.setdefault(_empower_base_name(upload_name), {})[extension] = (upload_name, payload)
            continue
        if extension and extension not in _EMPOWER_TEXT_EXPORT_EXTENSIONS:
            errors.append(
                f"{file.filename}: unsupported Empower import extension .{extension}; upload a CSV/ASCII export (.csv or .txt), Empower AIA .cdf, ARW chromatogram text, or ZIP containing .cdf/.arw files"
            )
            continue
        try:
            text = payload.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = payload.decode("latin-1", errors="ignore")
        lines = text.splitlines()
        rows = list(csv.DictReader(io.StringIO(text))) if lines and "," in lines[0] else []
        if rows:
            for row_idx, row in enumerate(rows, start=1):
                lower = {str(k).strip().lower(): v for k, v in row.items() if k is not None}
                try:
                    sample_name = _require_text(lower, ("sample", "sample_name", "sample name"), "Sample", row_idx, "Empower import")
                    sample_type = _require_text(lower, ("sample_type", "sample type"), "Sample Type", row_idx, "Empower import").upper()
                    injection_number = _require_text(lower, ("injection", "injection_number", "injection number"), "Injection Number", row_idx, "Empower import")
                    area = _require_finite(lower, ("area", "total area", "primary_peak_area"), "Area", row_idx, "Empower import")
                    rt = _require_finite(lower, ("rt", "retention_time", "retention time"), "Retention Time", row_idx, "Empower import")
                except HTTPException as exc:
                    errors.append(f"{file.filename} row {row_idx}: {exc.detail}")
                    continue
                primary_peak_percent = _finite_float(lower.get("primary_peak_percent") or lower.get("primary peak percent") or lower.get("area_percent") or lower.get("area percent"))
                record = {
                    "sample_name": sample_name,
                    "sample_type": sample_type,
                    "injection_number": injection_number,
                    "method_name": _first_text(lower, ("method", "method_name", "method name")),
                    "run_date": _first_text(lower, ("date", "run_date", "run date")),
                    "source_file": file.filename,
                    "source_format": "empower_csv_ascii_peak_table",
                    "total_area": area,
                    "primary_peak_area": area,
                    "primary_peak_percent": primary_peak_percent,
                    "primary_peak_rt": rt,
                    "primary_peak_resolution": _finite_float(lower.get("resolution")),
                    "peaks": [
                        {
                            "peak_id": 1,
                            "retention_time": rt,
                            "retention_time_min": rt,
                            "area": area,
                            "area_percent": primary_peak_percent,
                            "resolution": _finite_float(lower.get("resolution")),
                            "peak_source": "empower_csv_ascii_peak_table",
                        }
                    ],
                    "native_peak_count": 0,
                    "is_excluded": False,
                    "note": None,
                    "flag": None,
                    "import_engine": "csv",
                }
                injections.append(_finalize_empower_injection(record, import_id, injections))
        else:
            errors.append(f"{file.filename}: no CSV header row found; upload a real Empower CSV/ASCII export with sample, sample type, injection, area, and retention-time fields")
    if loose_cdf_arw:
        records, native_errors = _empower_records_from_grouped_cdf_arw(loose_cdf_arw, baseline_method, peak_prominence)
        errors.extend(native_errors)
        for record in records:
            injections.append(_finalize_empower_injection(record, import_id, injections))
    if not injections:
        detail = "; ".join(errors) if errors else "No real Empower injection rows parsed"
        raise HTTPException(status_code=400, detail=detail)
    source_format_counts: Dict[str, int] = {}
    for injection in injections:
        source_format = str(injection.get("source_format") or "unknown")
        source_format_counts[source_format] = source_format_counts.get(source_format, 0) + 1
    analytics = _empower_batch_analytics(injections, source_format_counts)
    response = {
        "import_id": import_id,
        "import_engine": "empower_cdf_arw_csv",
        "injections": injections,
        "n_injections": len(injections),
        "source_format_counts": source_format_counts,
        "sst_summary": _sst_summary(injections),
        "chromatogram_plotly_json": _empower_chromatogram_plot(injections),
        "qc_plotly_json": _empower_qc_plot(injections),
        "composition_plotly_json": _empower_composition_plot(analytics["peak_region_summary"]),
        "errors": errors,
        **analytics,
    }
    if persist:
        try:
            durable_ids = await persist_empower_import(
                import_session_id=import_id,
                source_files=source_files,
                injections=injections,
                analytics=response,
            )
        except _DB_DEGRADED_EXCEPTIONS as exc:
            _raise_db_service_degraded(exc)
        response.update(durable_ids)
        # The in-memory object is retained only as a short-lived review session cache;
        # durable chromatography rows are written to the analytical PostgreSQL store above.
        _EMPOWER_IMPORTS[import_id] = {
            "injections": injections,
            "errors": errors,
            "chromatogram_plotly_json": response["chromatogram_plotly_json"],
            "qc_plotly_json": response["qc_plotly_json"],
            "composition_plotly_json": response["composition_plotly_json"],
            "empower_summary": response["empower_summary"],
            "peak_table": response["peak_table"],
            "peak_region_summary": response["peak_region_summary"],
        }
    return _json_clean(response)

@router.get("/analysis/hplc/empower/sst")
def list_empower_sst(import_id: int) -> List[Dict[str, Any]]:
    record = _EMPOWER_IMPORTS.get(import_id)
    if not record:
        return []
    return _sst_summary(record["injections"])


@router.put("/analysis/hplc/empower/injections/{injection_id}")
def update_empower_injection(injection_id: int, payload: Dict[str, Any]) -> Dict[str, Any]:
    for record in _EMPOWER_IMPORTS.values():
        for inj in record["injections"]:
            if inj.get("id") == injection_id:
                inj.update({k: v for k, v in payload.items() if k in {"sample_name", "sample_type", "injection_number", "is_excluded", "note", "flag"}})
                return inj
    raise HTTPException(status_code=404, detail="Injection not found")


def _empower_csv(import_id: int) -> str:
    record = _EMPOWER_IMPORTS.get(import_id, {"injections": []})
    output = io.StringIO()
    fieldnames = ["id", "sample_name", "sample_type", "injection_number", "source_file", "total_area", "primary_peak_area", "primary_peak_percent", "primary_peak_rt", "primary_peak_resolution", "is_excluded", "flag", "note"]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row in record["injections"]:
        writer.writerow({k: row.get(k) for k in fieldnames})
    return output.getvalue()


@router.get("/analysis/hplc/empower/exports/sst-master")
def export_empower_sst_master(import_id: int) -> Response:
    return Response(_empower_csv(import_id), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=sst_master.csv"})


@router.get("/analysis/hplc/empower/exports/plasmid-tracking")
def export_empower_plasmid_tracking(import_id: int) -> Response:
    return Response(_empower_csv(import_id), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=plasmid_tracking.csv"})


class ControlChartRequest(BaseModel):
    data: List[float]
    subgroup_size: int = 1


class CapabilityRequest(BaseModel):
    data: List[float]
    usl: float
    lsl: float
    target: Optional[float] = None
    subgroup_size: int = 1


class DoEDesignRequest(BaseModel):
    design_type: str = "full_factorial"
    n_factors: int = 2
    center_points: int = 0


class RsmRequest(BaseModel):
    design_matrix: List[Dict[str, float]]
    response: List[float]


class HypothesisRequest(BaseModel):
    data: Optional[List[float]] = None
    pop_mean: Optional[float] = None
    group1: Optional[List[float]] = None
    group2: Optional[List[float]] = None
    before: Optional[List[float]] = None
    after: Optional[List[float]] = None
    groups: Optional[List[List[float]]] = None
    alpha: float = 0.05


class RegressionRequest(BaseModel):
    x: List[float]
    y: List[float]
    x_name: str = "X"
    y_name: str = "Y"


@router.post("/analysis/control-chart")
def control_chart(request: ControlChartRequest) -> Dict[str, Any]:
    data = [float(v) for v in request.data if math.isfinite(float(v))]
    if len(data) < 2:
        raise HTTPException(status_code=400, detail="At least two data points required")
    mean = _mean(data)
    sigma = _std(data)
    ucl = mean + 3 * sigma
    lcl = mean - 3 * sigma
    violations = [i for i, v in enumerate(data) if v > ucl or v < lcl]
    return _json_clean({"center_line": mean, "ucl": ucl, "lcl": lcl, "sigma": sigma, "violations": violations, "n": len(data), "plotly_json": {"data": [{"type": "scatter", "mode": "lines+markers", "y": data, "name": "Value"}], "layout": {"title": "Individuals Control Chart"}}})


@router.post("/analysis/capability")
def capability(request: CapabilityRequest) -> Dict[str, Any]:
    data = [float(v) for v in request.data if math.isfinite(float(v))]
    if len(data) < 2:
        raise HTTPException(status_code=400, detail="At least two data points required")
    mean = _mean(data)
    std_overall = _std(data)
    # Lightweight within-subgroup estimate: for individual data, use moving-range sigma;
    # for larger subgroups, pool within subgroup standard deviations when possible.
    subgroup_size = max(1, int(request.subgroup_size or 1))
    if subgroup_size <= 1 and len(data) >= 2:
        moving_ranges = [abs(b - a) for a, b in zip(data, data[1:])]
        std_within = _mean(moving_ranges) / 1.128 if moving_ranges else std_overall
    elif subgroup_size > 1:
        subgroup_stds = [_std(data[i : i + subgroup_size]) for i in range(0, len(data), subgroup_size) if len(data[i : i + subgroup_size]) > 1]
        std_within = _mean(subgroup_stds) if subgroup_stds else std_overall
    else:
        std_within = std_overall
    sigma = std_within or std_overall
    cp = (request.usl - request.lsl) / (6 * sigma) if sigma else None
    cpu = (request.usl - mean) / (3 * sigma) if sigma else None
    cpl = (mean - request.lsl) / (3 * sigma) if sigma else None
    cpk = min(v for v in [cpu, cpl] if v is not None) if sigma else None
    pp = (request.usl - request.lsl) / (6 * std_overall) if std_overall else None
    ppu = (request.usl - mean) / (3 * std_overall) if std_overall else None
    ppl = (mean - request.lsl) / (3 * std_overall) if std_overall else None
    ppk = min(v for v in [ppu, ppl] if v is not None) if std_overall else None
    nonconforming_count = sum(1 for v in data if v < request.lsl or v > request.usl)
    target = request.target if request.target is not None else (request.usl + request.lsl) / 2
    is_centered = abs(mean - target) <= max(sigma or 0.0, (request.usl - request.lsl) * 0.05)
    return _json_clean(
        {
            "mean": mean,
            "std_dev": sigma,
            "std_within": std_within,
            "std_overall": std_overall,
            "cp": cp,
            "cpu": cpu,
            "cpl": cpl,
            "cpk": cpk,
            "pp": pp,
            "ppu": ppu,
            "ppl": ppl,
            "ppk": ppk,
            "is_capable": bool(cpk is not None and cpk >= 1.33),
            "is_centered": bool(is_centered),
            "ppm_total": nonconforming_count / len(data) * 1_000_000,
            "n": len(data),
            "nonconforming_count": nonconforming_count,
            "plotly_json": {
                "data": [{"type": "histogram", "x": data, "name": "Observed"}],
                "layout": {
                    "title": "Process Capability",
                    "xaxis": {"title": "Value"},
                    "shapes": [
                        {"type": "line", "x0": request.lsl, "x1": request.lsl, "y0": 0, "y1": 1, "xref": "x", "yref": "paper", "line": {"color": "red", "dash": "dash"}},
                        {"type": "line", "x0": request.usl, "x1": request.usl, "y0": 0, "y1": 1, "xref": "x", "yref": "paper", "line": {"color": "red", "dash": "dash"}},
                        {"type": "line", "x0": target, "x1": target, "y0": 0, "y1": 1, "xref": "x", "yref": "paper", "line": {"color": "green", "dash": "dot"}},
                    ],
                },
            },
        }
    )


@router.post("/analysis/doe/design")
def doe_design(request: DoEDesignRequest) -> Dict[str, Any]:
    try:
        import pyDOE3
    except ImportError as exc:  # pragma: no cover - dependency is required by pyproject
        raise HTTPException(status_code=500, detail="pyDOE3 is not installed in the BMS API environment") from exc

    n = max(1, min(int(request.n_factors), 8))
    factor_names = [f"X{i + 1}" for i in range(n)]
    dtype = request.design_type.lower()
    center_points = max(0, int(request.center_points))
    generator = "pyDOE3.ff2n"

    if dtype == "central_composite":
        matrix = pyDOE3.ccdesign(n, center=(center_points, center_points), alpha="orthogonal", face="circumscribed")
        generator = "pyDOE3.ccdesign"
    elif dtype == "box_behnken":
        if n < 3:
            matrix = pyDOE3.ff2n(n)
            generator = "pyDOE3.ff2n"
        else:
            matrix = pyDOE3.bbdesign(n, center=center_points)
            generator = "pyDOE3.bbdesign"
    elif dtype == "plackett_burman":
        matrix = pyDOE3.pbdesign(n)
        generator = "pyDOE3.pbdesign"
        if center_points:
            matrix = np.vstack([matrix, np.zeros((center_points, n))])
    elif dtype == "fractional_factorial" and n >= 3:
        base = pyDOE3.ff2n(n - 1)
        generated = []
        for row in base:
            last = float(np.prod(row))
            generated.append([*row.tolist(), last])
        matrix = np.asarray(generated, dtype=float)
        generator = "pyDOE3.ff2n(generator=product)"
        if center_points:
            matrix = np.vstack([matrix, np.zeros((center_points, n))])
    else:
        matrix = pyDOE3.ff2n(n)
        generator = "pyDOE3.ff2n"
        if center_points:
            matrix = np.vstack([matrix, np.zeros((center_points, n))])

    runs = [dict(zip(factor_names, [float(v) for v in row])) for row in np.asarray(matrix, dtype=float).tolist()]
    factors = [{"name": name, "low": -1, "center": 0, "high": 1, "type": "continuous"} for name in factor_names]
    return _json_clean(
        {
            "analysis_engine": "pyDOE3",
            "engine_package": "pyDOE3",
            "design_type": request.design_type,
            "n_runs": len(runs),
            "n_factors": n,
            "factors": factors,
            "design_matrix": runs,
            "summary": f"{request.design_type} DOE generated with {generator}: {len(runs)} runs, {n} factors, and {center_points} requested center points.",
            "metadata": {
                "analysis_family": "JMP-compatible DOE",
                "design_generator": generator,
                "supported_models": ["main effects", "two-factor interactions", "quadratic/RSM terms", "desirability optimization"],
            },
        }
    )


@router.post("/analysis/doe/rsm")
def rsm_analysis(request: RsmRequest) -> Dict[str, Any]:
    try:
        import statsmodels.api as sm
    except ImportError as exc:  # pragma: no cover - dependency is required by pyproject
        raise HTTPException(status_code=500, detail="statsmodels is not installed in the BMS API environment") from exc

    if len(request.design_matrix) != len(request.response) or not request.design_matrix:
        raise HTTPException(status_code=400, detail="Design matrix rows must match response values")
    cols = list(request.design_matrix[0].keys())
    x_rows = []
    for row_idx, row in enumerate(request.design_matrix, start=1):
        values = []
        for c in cols:
            if c not in row:
                raise HTTPException(status_code=400, detail=f"Design matrix row {row_idx} missing factor {c}")
            finite = _finite_float(row[c])
            if finite is None:
                raise HTTPException(status_code=400, detail=f"Design matrix row {row_idx} factor {c} must be finite")
            values.append(finite)
        x_rows.append(values)
    x = np.asarray(x_rows, dtype=float)
    y = np.asarray(request.response, dtype=float)
    if not np.all(np.isfinite(x)) or not np.all(np.isfinite(y)):
        raise HTTPException(status_code=400, detail="RSM inputs must be finite numeric values")

    term_arrays = [np.ones(len(x))]
    term_names = ["Intercept"]
    for i, c in enumerate(cols):
        term_arrays.append(x[:, i])
        term_names.append(c)
    for i, c in enumerate(cols):
        term_arrays.append(x[:, i] ** 2)
        term_names.append(f"{c}^2")
    for i in range(len(cols)):
        for j in range(i + 1, len(cols)):
            term_arrays.append(x[:, i] * x[:, j])
            term_names.append(f"{cols[i]}:{cols[j]}")

    X = np.vstack(term_arrays).T
    model = sm.OLS(y, X).fit()
    beta = np.asarray(model.params, dtype=float)
    coeffs = dict(zip(term_names, beta.tolist()))

    linear = np.asarray([coeffs.get(c, 0.0) for c in cols], dtype=float)
    hessian = np.zeros((len(cols), len(cols)), dtype=float)
    for i, c in enumerate(cols):
        hessian[i, i] = 2.0 * coeffs.get(f"{c}^2", 0.0)
    for i in range(len(cols)):
        for j in range(i + 1, len(cols)):
            value = coeffs.get(f"{cols[i]}:{cols[j]}", 0.0)
            hessian[i, j] = value
            hessian[j, i] = value
    stationary_status = "estimated"
    optimal: Optional[Dict[str, float]] = None
    predicted_optimum: Optional[float] = None
    try:
        optimum_vec = -np.linalg.solve(hessian, linear)
        if not np.all(np.isfinite(optimum_vec)):
            raise np.linalg.LinAlgError("non-finite stationary point")
        optimum_vec = np.clip(optimum_vec, -2.0, 2.0)
        optimal = {c: float(optimum_vec[i]) for i, c in enumerate(cols)}
    except np.linalg.LinAlgError:
        stationary_status = "not_estimable_singular_hessian"
        optimum_vec = None

    def predict_row(values: np.ndarray) -> float:
        terms = [1.0]
        terms.extend(values.tolist())
        terms.extend((values**2).tolist())
        for i in range(len(cols)):
            for j in range(i + 1, len(cols)):
                terms.append(float(values[i] * values[j]))
        return float(np.dot(np.asarray(terms, dtype=float), beta))

    if optimum_vec is not None:
        predicted_optimum = predict_row(optimum_vec)
    p_values = {name: (float(p) if math.isfinite(float(p)) else None) for name, p in zip(term_names, model.pvalues)}

    contour_plot: Optional[Dict[str, Any]] = None
    surface_plot: Optional[Dict[str, Any]] = None
    if len(cols) >= 2:
        x1_grid = np.linspace(-1.5, 1.5, 31)
        x2_grid = np.linspace(-1.5, 1.5, 31)
        plot_anchor = optimum_vec.copy() if optimum_vec is not None else np.zeros(len(cols), dtype=float)
        z_grid = []
        for yv in x2_grid:
            row_values = []
            for xv in x1_grid:
                values = plot_anchor.copy()
                values[0] = xv
                values[1] = yv
                row_values.append(predict_row(values))
            z_grid.append(row_values)
        contour_plot = {
            "data": [{"type": "contour", "x": x1_grid.tolist(), "y": x2_grid.tolist(), "z": z_grid, "colorscale": "Viridis"}],
            "layout": {"title": "RSM Contour", "xaxis": {"title": cols[0]}, "yaxis": {"title": cols[1]}},
        }
        surface_plot = {
            "data": [{"type": "surface", "x": x1_grid.tolist(), "y": x2_grid.tolist(), "z": z_grid, "colorscale": "Viridis"}],
            "layout": {"title": "RSM Surface", "scene": {"xaxis": {"title": cols[0]}, "yaxis": {"title": cols[1]}, "zaxis": {"title": "Predicted response"}}},
        }

    return _json_clean(
        {
            "analysis_engine": "statsmodels.OLS",
            "engine_package": "statsmodels",
            "summary": "Response-surface model fit with statsmodels OLS: intercept, linear, quadratic, and two-factor interaction terms.",
            "terms": term_names,
            "coefficients": coeffs,
            "p_values": p_values,
            "r_squared": float(model.rsquared) if math.isfinite(float(model.rsquared)) else None,
            "adj_r_squared": float(model.rsquared_adj) if math.isfinite(float(model.rsquared_adj)) else None,
            "aic": float(model.aic) if math.isfinite(float(model.aic)) else None,
            "bic": float(model.bic) if math.isfinite(float(model.bic)) else None,
            "optimal_point": optimal,
            "stationary_point_status": stationary_status,
            "predicted_optimum": predicted_optimum,
            "contour_plot": contour_plot,
            "surface_plot": surface_plot,
        }
    )


@router.post("/analysis/hypothesis/t-test/one-sample")
def hypothesis_one_sample(request: HypothesisRequest) -> Dict[str, Any]:
    data = request.data or []
    if request.pop_mean is None or len(data) < 2:
        raise HTTPException(status_code=400, detail="data and pop_mean required")
    t, p = stats.ttest_1samp(data, request.pop_mean)
    return _json_clean({"test": "one_sample_t", "statistic": t, "p_value": p, "significant": p < request.alpha, "mean": _mean(data), "alpha": request.alpha})


@router.post("/analysis/hypothesis/t-test/two-sample")
def hypothesis_two_sample(request: HypothesisRequest) -> Dict[str, Any]:
    if not request.group1 or not request.group2:
        raise HTTPException(status_code=400, detail="group1 and group2 required")
    t, p = stats.ttest_ind(request.group1, request.group2, equal_var=False)
    return _json_clean({"test": "welch_t", "statistic": t, "p_value": p, "significant": p < request.alpha, "mean_group1": _mean(request.group1), "mean_group2": _mean(request.group2), "alpha": request.alpha})


@router.post("/analysis/hypothesis/t-test/paired")
def hypothesis_paired(request: HypothesisRequest) -> Dict[str, Any]:
    if not request.before or not request.after or len(request.before) != len(request.after):
        raise HTTPException(status_code=400, detail="paired before/after vectors of equal length required")
    t, p = stats.ttest_rel(request.before, request.after)
    return _json_clean({"test": "paired_t", "statistic": t, "p_value": p, "significant": p < request.alpha, "alpha": request.alpha})


@router.post("/analysis/hypothesis/anova")
def hypothesis_anova(request: HypothesisRequest) -> Dict[str, Any]:
    groups = request.groups or []
    if len(groups) < 2:
        raise HTTPException(status_code=400, detail="At least two groups required")
    f, p = stats.f_oneway(*groups)
    return _json_clean({"test": "one_way_anova", "statistic": f, "p_value": p, "significant": p < request.alpha, "alpha": request.alpha})


@router.post("/analysis/regression/simple")
def regression_simple(request: RegressionRequest) -> Dict[str, Any]:
    try:
        import statsmodels.api as sm
    except ImportError as exc:  # pragma: no cover - dependency is required by pyproject
        raise HTTPException(status_code=500, detail="statsmodels is not installed in the BMS API environment") from exc

    x = np.asarray([float(v) for v in request.x], dtype=float)
    y = np.asarray([float(v) for v in request.y], dtype=float)
    if len(x) != len(y) or len(x) < 2:
        raise HTTPException(status_code=400, detail="At least two paired finite points are required")
    if not np.all(np.isfinite(x)) or not np.all(np.isfinite(y)):
        raise HTTPException(status_code=400, detail="Regression inputs must be finite numeric values")

    X = sm.add_constant(x, has_constant="add")
    model = sm.OLS(y, X).fit()
    intercept = float(model.params[0])
    slope = float(model.params[1])
    fitted = model.fittedvalues.astype(float).tolist()
    residuals = model.resid.astype(float).tolist()
    r2 = float(model.rsquared) if math.isfinite(float(model.rsquared)) else None
    adj_r2 = float(model.rsquared_adj) if math.isfinite(float(model.rsquared_adj)) else r2
    f_stat = float(model.fvalue) if model.fvalue is not None and math.isfinite(float(model.fvalue)) else float("inf")
    f_pvalue = float(model.f_pvalue) if model.f_pvalue is not None and math.isfinite(float(model.f_pvalue)) else 0.0
    slope_p = float(model.pvalues[1]) if len(model.pvalues) > 1 and math.isfinite(float(model.pvalues[1])) else None
    intercept_p = float(model.pvalues[0]) if len(model.pvalues) > 0 and math.isfinite(float(model.pvalues[0])) else None
    n = int(model.nobs)
    equation = f"{request.y_name} = {intercept:.4g} + {slope:.4g}*{request.x_name}"
    return _json_clean(
        {
            "analysis_engine": "statsmodels.OLS",
            "engine_package": "statsmodels",
            "slope": slope,
            "intercept": intercept,
            "r_squared": r2,
            "p_value": slope_p,
            "std_err": float(model.bse[1]) if len(model.bse) > 1 and math.isfinite(float(model.bse[1])) else None,
            "x_name": request.x_name,
            "y_name": request.y_name,
            "equation": equation,
            "coefficients": {"intercept": intercept, request.x_name: slope, "X": slope},
            "adj_r_squared": adj_r2,
            "f_statistic": f_stat,
            "f_pvalue": f_pvalue,
            "aic": float(model.aic) if math.isfinite(float(model.aic)) else None,
            "bic": float(model.bic) if math.isfinite(float(model.bic)) else None,
            "n_obs": n,
            "p_values": {"intercept": intercept_p, request.x_name: slope_p, "X": slope_p},
            "scatter_plot": {
                "data": [
                    {"type": "scatter", "mode": "markers", "x": x.tolist(), "y": y.tolist(), "name": "Observed"},
                    {"type": "scatter", "mode": "lines", "x": x.tolist(), "y": fitted, "name": "Fit"},
                ],
                "layout": {"title": "Simple Linear Regression", "xaxis": {"title": request.x_name}, "yaxis": {"title": request.y_name}},
            },
            "diagnostics_plot": {
                "data": [{"type": "scatter", "mode": "markers", "x": fitted, "y": residuals, "name": "Residuals"}],
                "layout": {"title": "Residuals vs Fitted", "xaxis": {"title": "Fitted"}, "yaxis": {"title": "Residual"}},
            },
        }
    )


@router.get("/tools")
def external_tools() -> Dict[str, Any]:
    return _json_clean({"tools": assay_tool_registry(), "by_category": tools_by_category()})


@router.get("/analytical-store/status")
def analytical_store() -> Dict[str, Any]:
    return _json_clean(analytical_store_status())


@router.get("/capabilities")
def capabilities() -> Dict[str, Any]:
    tools = assay_tool_registry()
    return _json_clean(
        {
            "service": "BioModStack Stats Toolkit",
            "surfaces": ["qPCR QuantStudio/StepOnePlus", "Waters/Empower chromatography", "plasmid DNA isoform analysis", "JMP-like DOE/statistics", "Plotly visualization"],
            "source_of_truth": "BMS API /api/assay-analytics",
            "not_used": "legacy standalone parser service",
            "analytical_store": analytical_store_status(),
            "external_tools": tools,
            "external_tools_by_category": tools_by_category(),
        }
    )
